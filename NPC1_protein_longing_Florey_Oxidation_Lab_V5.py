# -*- coding: utf-8 -*-
"""
Created on Sun Feb 23 19:16:45 2025
@author: Mr. Lilian LAGARRIGUE
@email: lilian.lagarrigue@hotmail.com
VERSION 1 : Created on Tue Jan 21 12:51:54 2025 - ERROR IN THE EQUATIONS
VERSION 2 : Created on Mon Jan 27 18:48:53 2025 - CORRECTION OF THE ERROR IN THE EQUATIONS
VERSION 3 : Created on Wed Jan 29 16:08:30 2025 - Rescaling the protein concentration experimental curve depending on alpha_protein ; beta_protein ; alpha_mRNA; beta_mRNA
VERSION 4 : Created on Sat Mar 01 10:00:00 2025 - Updating the search loop method. Adding some intermediate plots. Adding the parameters and MSE evolution to the plot and in the command window. Cleaning the code.
VERSION 5 : Created on Wed Apr 10 13:10:00 2025 - 

Save an additional excel file containing: [ONGOING]
    the mRNA dosage
    the associated half-life time
    the curve fitting parameters
Plot the predicted mRNA concentration. [ONGOING]
Measure the predicted mRNA half-life based on the predicted mRNA concentration. [ONGOING]
"""

#############################################################
## CLEAR CONSOLE ##                                         #
import os                                                   #
from IPython import get_ipython                             #
try: os.system('cls' if os.name == 'nt' else 'clear')       #
except:                                                     #
    try: get_ipython().magic("clear") # Clears the console  #
    except: pass                                            #
#############################################################

# Parameters on the expected structure of the CSV file
Nb_Data_Per_Sheet               = 3
Data_Labels_Row                 = 0
Time_Labels_Row                 = 1
mRNA_Dosages_Row                = 1
Time_Row_START                  = 2
Protein_Concentration_Row_START = 2
Data_Labels_Column_OFFSET       = 0
Time_Labels_Column_OFFSET       = 0
mRNA_Dosages_Column_OFFSET      = 1
Nb_Columns_Per_Data             = 2

# Boolean parameter indicating whether to replace the non-breaking space character by a single space caracter (as ' ') or not
BOOL_ReplaceNonBreakingSpace_by_space = True

# Estimated number of NPC1 proteins at equilibrium (right before the mRNA injection)
Number_of_NPC1_proteins_before_injection_at_equilibrium_MIN_ESTIMATED = 500
Number_of_NPC1_proteins_before_injection_at_equilibrium_____ESTIMATED = 1000
Number_of_NPC1_proteins_before_injection_at_equilibrium_MAX_ESTIMATED = 2000
Number_of_NPC1_proteins_before_injection_at_peak_MAX_ESTIMATED        = 20000

coef_multiplicator_between_1_and_2_for_estimated_MAX = 2 # coefficient multiplying the maximum value taken for the experimental data scaling so we allow the fitting curve to have a slightly higher maximum value ...

# # # # # # # # # # # # # # # # # # # # #
## Parameters on the normalization step
NORM_VAL_MIN = 0 # Lowest value after normalization
NORM_VAL_MAX = 1 # Highest value after normalization
MIN_VALUE_FOR_NORMALIZED_FIT = 0
MAX_VALUE_FOR_NORMALIZED_FIT = 1

# Boolean telling whether to scale the experimental data and fitting curve on the same scale
BOOL_plot_intermediate_curves_with_EXP_DATA_and_FIT = False

# Parameters on the process progress message
ShowProgression_PercentageStep = 1
round_digit_ShowProgression    = 0

# Choosing the method to generate the parameters grid search
METHOD_GENERATE_PARAMETERS_GRID_SEARCH = "Power_min_to_power_max" # "Power_min_to_power_max" # "Powers_in_a_given_list"
                                        # "Power_min_to_power_max" means the 10th powers min and max values are taken as the lower and upper bonds for the searching grid
                                        # "Powers_in_a_given_list" means the 10th powers can be discontinuous such as 1e-9 to 1e-7 and the 1e-3 and 1e+5 etc ...

# Parameters on the search grid
Nb_points_per_decade = 8 # NEVER LOWER THAN 1 (can be higher than 9 because the points are generated via the np.linspace function)
                          # If only one or two decimal integers required for the linearspace, chose the 'Nb_points_per_decade' in this [2, 3, 5, 6, 9, 11, 17, 21, 26, 33, 41, 51 ... to complete if needed ...]
PowerStep_ALL = 1

# Parameters for the grid search of every different alpha and beta
Nb_points_per_decade_AP = Nb_points_per_decade  # SET
MIN_Power_AP            = -16
MAX_Power_AP            = -1
PowerStep_AP            = PowerStep_ALL         # SET

Nb_points_per_decade_BP = Nb_points_per_decade  # SET
MIN_Power_BP            = -13
MAX_Power_BP            = 0                     # SET
PowerStep_BP            = PowerStep_ALL         # SET

Nb_points_per_decade_Am = Nb_points_per_decade  # SET
MIN_Power_Am            = -7
MAX_Power_Am            = +6
PowerStep_Am            = PowerStep_ALL         # SET

Nb_points_per_decade_Bm = Nb_points_per_decade  # SET
MIN_Power_Bm            = MIN_Power_BP          # SET
MAX_Power_Bm            = abs(MIN_Power_BP)     # SET
PowerStep_Bm            = PowerStep_ALL         # SET

# ###############################################################################
# ###############################################################################
# ## THIS BLOCK CAN BE USED TO TUNE THE PARAMETERS RANGES WHEN ROUGHLY IDENTIFIED
# # Parameters for the grid search of every different alpha and beta
# Nb_points_per_decade    = 50
# Nb_points_per_decade_AP = Nb_points_per_decade  # SET
# MIN_Power_AP            = -10
# MAX_Power_AP            = -8
# PowerStep_AP            = PowerStep_ALL         # SET
# Nb_points_per_decade_BP = Nb_points_per_decade  # SET
# MIN_Power_BP            = -6
# MAX_Power_BP            = -4 # 0                     # SET
# PowerStep_BP            = PowerStep_ALL         # SET
# Nb_points_per_decade_Am = Nb_points_per_decade  # SET
# MIN_Power_Am            = +0
# MAX_Power_Am            = +3
# PowerStep_Am            = PowerStep_ALL         # SET
# Nb_points_per_decade_Bm = Nb_points_per_decade  # SET
# MIN_Power_Bm            = MIN_Power_BP          # SET
# MAX_Power_Bm            = abs(MIN_Power_BP)     # SET
# PowerStep_Bm            = PowerStep_ALL         # SET
# ## THIS BLOCK CAN BE USED TO TUNE THE PARAMETERS RANGES WHEN ROUGHLY IDENTIFIED
# ###############################################################################
# ###############################################################################

import csv
import math
import matplotlib.pyplot as plt
import numpy             as np
import openpyxl
import os
import re
# import sys
import time

from datetime import datetime
from scipy.constants import Avogadro

# Getting the current date and time
DATE_AND_TIME = datetime.now().strftime("%Y_%m_%d_%H_%M_%S") # Format Year_Month_Day_Hour_Minute_Second

# - Function taking a mRNA sequence in argument and giving the number of synonymous mRNA sequences + the protein sequence expected after translation process
def CountSynonymousmRNASequences_from_mRNA(mRNA_sequence , FORMAT = 'TEXT' , mRNASeq_StartingWord = '' , PrintMessage = False , ShowCombinationEvolutionPlot = False , ORDER_TYPE = 'SORT_by_%'):
    AminoAcidToCodon = {                            'Alanine'       : [4 , 'GCU', 'GCC', 'GCA', 'GCG']              , 'Ala' : [4 , 'GCU', 'GCC', 'GCA', 'GCG']              , 'A' : [4 , 'GCU', 'GCC', 'GCA', 'GCG'],
                                                    'Valine'        : [4 , 'GUU', 'GUC', 'GUA', 'GUG']              , 'Val' : [4 , 'GUU', 'GUC', 'GUA', 'GUG']              , 'V' : [4 , 'GUU', 'GUC', 'GUA', 'GUG'] ,
                                                    'Leucine'       : [6 , 'UUA', 'UUG', 'CUU', 'CUC', 'CUA', 'CUG'], 'Leu' : [6 , 'UUA', 'UUG', 'CUU', 'CUC', 'CUA', 'CUG'], 'L' : [6 , 'UUA', 'UUG', 'CUU', 'CUC', 'CUA', 'CUG'] ,
                                                    'Isoleucine'    : [3 , 'AUU', 'AUC', 'AUA']                     , 'Ile' : [3 , 'AUU', 'AUC', 'AUA']                     , 'I' : [3 , 'AUU', 'AUC', 'AUA'] ,
        'Start_Codon'   : [1 , 'AUG']             , 'Methionine'    : [1 , 'AUG']                                   , 'Met' : [1 , 'AUG']                                   , 'M' : [1 , 'AUG'] ,
                                                    'Phenylalanine' : [2 , 'UUU', 'UUC']                            , 'Phe' : [2 , 'UUU', 'UUC']                            , 'F' : [2 , 'UUU', 'UUC'] ,
                                                    'Tryptophan'    : [1 , 'UGG']                                   , 'Trp' : [1 , 'UGG']                                   , 'W' : [1 , 'UGG'] ,
                                                    'Proline'       : [4 , 'CCU', 'CCC', 'CCA', 'CCG']              , 'Pro' : [4 , 'CCU', 'CCC', 'CCA', 'CCG']              , 'P' : [4 , 'CCU', 'CCC', 'CCA', 'CCG'] ,
                                                    'Serine'        : [6 , 'UCU', 'UCC', 'UCA', 'UCG', 'AGU', 'AGC'], 'Ser' : [6 , 'UCU', 'UCC', 'UCA', 'UCG', 'AGU', 'AGC'], 'S' : [6 , 'UCU', 'UCC', 'UCA', 'UCG', 'AGU', 'AGC'] ,
                                                    'Threonine'     : [4 , 'ACU', 'ACC', 'ACA', 'ACG']              , 'Thr' : [4 , 'ACU', 'ACC', 'ACA', 'ACG']              , 'T' : [4 , 'ACU', 'ACC', 'ACA', 'ACG'] ,
                                                    'Cysteine'      : [2 , 'UGU', 'UGC']                            , 'Cys' : [2 , 'UGU', 'UGC']                            , 'C' : [2 , 'UGU', 'UGC'] ,
                                                    'Tyrosine'      : [2 , 'UAU', 'UAC']                            , 'Tyr' : [2 , 'UAU', 'UAC']                            , 'Y' : [2 , 'UAU', 'UAC'] ,
                                                    'Asparagine'    : [2 , 'AAU', 'AAC']                            , 'Asn' : [2 , 'AAU', 'AAC']                            , 'N' : [2 , 'AAU', 'AAC'] ,
                                                    'Glutamine'     : [2 , 'CAA', 'CAG']                            , 'Gln' : [2 , 'CAA', 'CAG']                            , 'Q' : [2 , 'CAA', 'CAG'] ,
                                                    'Lysine'        : [2 , 'AAA', 'AAG']                            , 'Lys' : [2 , 'AAA', 'AAG']                            , 'K' : [2 , 'AAA', 'AAG'] ,
                                                    'Arginine'      : [6 , 'CGU', 'CGC', 'CGA', 'CGG', 'AGA', 'AGG'], 'Arg' : [6 , 'CGU', 'CGC', 'CGA', 'CGG', 'AGA', 'AGG'], 'R' : [6 , 'CGU', 'CGC', 'CGA', 'CGG', 'AGA', 'AGG'] ,
                                                    'Histidine'     : [2 , 'CAU', 'CAC']                            , 'His' : [2 , 'CAU', 'CAC']                            , 'H' : [2 , 'CAU', 'CAC'] ,
                                                    'Aspartic_acid' : [2 , 'GAU', 'GAC']                            , 'Asp' : [2 , 'GAU', 'GAC']                            , 'D' : [2 , 'GAU', 'GAC'] ,
                                                    'Glutamic_acid' : [2 , 'GAA', 'GAG']                            , 'Glu' : [2 , 'GAA', 'GAG']                            , 'E' : [2 , 'GAA', 'GAG'] ,
                                                    'Glycine'       : [4 , 'GGU', 'GGC', 'GGA', 'GGG']              , 'Gly' : [4 , 'GGU', 'GGC', 'GGA', 'GGG']              , 'G' : [4 , 'GGU', 'GGC', 'GGA', 'GGG'] ,
        'Stop_Codons'   : [3 , 'UAA', 'UAG', 'UGA']                                                                                                                         , 'X' : [3 , 'UAA', 'UAG', 'UGA']}

    # Creating the inverted dictionary
    CodonToAminoAcid = {}
    for key , value in AminoAcidToCodon.items() :
        for codon in value[1:] :
            if codon not in CodonToAminoAcid : CodonToAminoAcid[codon] = []
            CodonToAminoAcid[codon].append(key)
    
    # Initializing the different parameters / constants / variables
    string_length             = len(mRNA_sequence)
    mRNA_seq_FILTERED         = ''
    mRNASequenceStartingIndex = 0 # by default we initialize the protein starting caracter position as the first one in the whole string
    total_combinations        = 1 # Initialize total combinations
    ignored_caracters         = 0
    Dict_CodonCounts = {'GCU': [0 , 0 , []], # The two zeros represent the initial counts and % of apparition of each codon in the sequence relatively to all the possible codons
                        'GCC': [0 , 0 , []],
                        'GCA': [0 , 0 , []],
                        'GCG': [0 , 0 , []],
                        'GUU': [0 , 0 , []],
                        'GUC': [0 , 0 , []],
                        'GUA': [0 , 0 , []],
                        'GUG': [0 , 0 , []],
                        'UUA': [0 , 0 , []],
                        'UUG': [0 , 0 , []],
                        'CUU': [0 , 0 , []],
                        'CUC': [0 , 0 , []],
                        'CUA': [0 , 0 , []],
                        'CUG': [0 , 0 , []],
                        'AUU': [0 , 0 , []],
                        'AUC': [0 , 0 , []],
                        'AUA': [0 , 0 , []],
                        'AUG': [0 , 0 , []],
                        'UUU': [0 , 0 , []],
                        'UUC': [0 , 0 , []],
                        'UGG': [0 , 0 , []],
                        'CCU': [0 , 0 , []],
                        'CCC': [0 , 0 , []],
                        'CCA': [0 , 0 , []],
                        'CCG': [0 , 0 , []],
                        'UCU': [0 , 0 , []],
                        'UCC': [0 , 0 , []],
                        'UCA': [0 , 0 , []],
                        'UCG': [0 , 0 , []],
                        'AGU': [0 , 0 , []],
                        'AGC': [0 , 0 , []],
                        'ACU': [0 , 0 , []],
                        'ACC': [0 , 0 , []],
                        'ACA': [0 , 0 , []],
                        'ACG': [0 , 0 , []],
                        'UGU': [0 , 0 , []],
                        'UGC': [0 , 0 , []],
                        'UAU': [0 , 0 , []],
                        'UAC': [0 , 0 , []],
                        'AAU': [0 , 0 , []],
                        'AAC': [0 , 0 , []],
                        'CAA': [0 , 0 , []],
                        'CAG': [0 , 0 , []],
                        'AAA': [0 , 0 , []],
                        'AAG': [0 , 0 , []],
                        'CGU': [0 , 0 , []],
                        'CGC': [0 , 0 , []],
                        'CGA': [0 , 0 , []],
                        'CGG': [0 , 0 , []],
                        'AGA': [0 , 0 , []],
                        'AGG': [0 , 0 , []],
                        'CAU': [0 , 0 , []],
                        'CAC': [0 , 0 , []],
                        'GAU': [0 , 0 , []],
                        'GAC': [0 , 0 , []],
                        'GAA': [0 , 0 , []],
                        'GAG': [0 , 0 , []],
                        'GGU': [0 , 0 , []],
                        'GGC': [0 , 0 , []],
                        'GGA': [0 , 0 , []],
                        'GGG': [0 , 0 , []],
                        'UAA': [0 , 0 , []],
                        'UAG': [0 , 0 , []],
                        'UGA': [0 , 0 , []]}
    
    # Processing the sequence depending on the file format
    if FORMAT == 'TEXT': mRNASequenceStartingIndex = 0
    elif FORMAT == 'FASTA':
        mRNASequenceStartingIndex = mRNA_sequence.rfind(mRNASeq_StartingWord)
        if mRNASequenceStartingIndex == -1:
            mRNASequenceStartingIndex = 0
            print(f"\nWARNING : The mRNA sequence supposed to be in '{FORMAT}' format does not contain any required starting-sequence-word such as '{mRNASeq_StartingWord}'. Index 0 will be considered as the starting index for the mRNA sequence.")
        else: mRNASequenceStartingIndex = mRNASequenceStartingIndex + len(mRNASeq_StartingWord) # the starting index begins just after the starting word in that case.
    
    # Counts every nucleotide occurrence
    A_NucleotideCount = mRNA_sequence[mRNASequenceStartingIndex : string_length].count('A')
    C_NucleotideCount = mRNA_sequence[mRNASequenceStartingIndex : string_length].count('C')
    G_NucleotideCount = mRNA_sequence[mRNASequenceStartingIndex : string_length].count('G')
    T_NucleotideCount = mRNA_sequence[mRNASequenceStartingIndex : string_length].count('T')
    U_NucleotideCount = mRNA_sequence[mRNASequenceStartingIndex : string_length].count('U')
    NbNucleotides_TOT = A_NucleotideCount + C_NucleotideCount + G_NucleotideCount + T_NucleotideCount + U_NucleotideCount
    A_NucleotidePercentage = round(100 * A_NucleotideCount / NbNucleotides_TOT , 2)
    C_NucleotidePercentage = round(100 * C_NucleotideCount / NbNucleotides_TOT , 2)
    G_NucleotidePercentage = round(100 * G_NucleotideCount / NbNucleotides_TOT , 2)
    T_NucleotidePercentage = round(100 * T_NucleotideCount / NbNucleotides_TOT , 2)
    U_NucleotidePercentage = round(100 * U_NucleotideCount / NbNucleotides_TOT , 2)
    
    # Iterate over each codon in the protein sequence
    reading_offset_tmp = 0
    for i in range(mRNASequenceStartingIndex , string_length-3 , 3):
        
        if string_length-1 < i+3+reading_offset_tmp and (i+3+reading_offset_tmp - (string_length-1)) < 3:    
            # print(f"\nWARNING: the last {string_length-1 - (i+reading_offset_tmp)} caracter(s) '{mRNA_sequence[i+reading_offset_tmp:-1]}' might have been ignored at the end of the sequence due to necessary reading offset...")
            for j , residual_caracters in enumerate(mRNA_sequence[i+reading_offset_tmp:-1]):
                if residual_caracters in "ACGTU" : mRNA_seq_FILTERED += residual_caracters
                else: ignored_caracters += 1
        elif string_length-1 < i+3+reading_offset_tmp and (i+3+reading_offset_tmp - (string_length-1)) >= 3: pass # In anycase the end of the sequence is reached, the for loop ignores the remaining 'i'th steps in the loop because all the sequence has already been treated ...
        else:
            codon = mRNA_sequence[reading_offset_tmp+i:i+3+reading_offset_tmp]
            if '\n' in codon:
                codon = mRNA_sequence[reading_offset_tmp+i:i+3+reading_offset_tmp+1]
                reading_offset_tmp += 1 # THIS LINE AFTER HAVING READ THE CODON WITH ONE MORE CARACTER (see the '+1' in the upper line ...)
                ignored_caracters  += 1
                codon = codon.replace('\n', '')
            codon = codon.replace('T', 'U') # This line to use the same table with Uracile nucleotide for codon identification instead of Thymine ...
            if codon not in CodonToAminoAcid:
                print(f"Invalid codon (position #{i}): '{codon}' is not known as a codon.")
                ignored_caracters = ignored_caracters + 3
            else:
                # Adds +1 to the corresponding codon count for statistics
                Dict_CodonCounts[codon][0] += 1
                
                # Adds the codon string chain to the mRNA sequence cleaned
                mRNA_seq_FILTERED += codon
                
                # Adds the index of the current codon found to its list of positions
                Dict_CodonCounts[codon][2].append(i) # simply adds 'i' because whatever the ignored caracters and offset etc, the ith position still describes the current codon that is read from the begining to the end ...
                
                # Nucleotidei , Nucleotideii , Nucleotideiii = codon
                amino_acid_name                            = CodonToAminoAcid[codon][0] #index 0 for full name , 1 for 3letters name , 2 for 1 letter name
                amino_acid_combinations                    = len(AminoAcidToCodon[amino_acid_name])
                total_combinations                        *= amino_acid_combinations
    
    # Computing the main results
    mRNA_seq_length                                            = len(mRNA_seq_FILTERED) # string_length - mRNASequenceStartingIndex - ignored_caracters # this is the correctly scaled mRNA sequence length only
    Ncodons                                                    = mRNA_seq_length // 3
    total_combinations_EXPONENTIAL_PART_BASE_10                = int(math.log10(total_combinations)) # The total combination number written as 'T' is decomposed as T = aT * 10**(bT) with aT = 1 and bT = log10(T)
    AverageDegenerationDegreePerCodon_EXPONENTIAL_PART_BASE_10 = total_combinations_EXPONENTIAL_PART_BASE_10 / Ncodons # The total combination number written as 'T' is decomposed as T = aT * 10**(bT) with aT = 1 and bT = log10(T) thus, T**(1/L) = (aT*10**bT)**(1/L) = 10**(bT/L)
    AverageDegenerationDegreePerCodon                          = pow(10 , AverageDegenerationDegreePerCodon_EXPONENTIAL_PART_BASE_10)
    
    # Filling the % of apparition of every codon in the dictionnary
    for codon_keys in Dict_CodonCounts: Dict_CodonCounts[codon_keys][1] = Dict_CodonCounts[codon_keys][0] * 100 / Ncodons
    
    # Eventually print the output messages
    if PrintMessage == True :
        
        # Printing the information about the mRNA sequence reading part
        print("\t------------------")
        print(f"\nmRNA sequence '{mRNA_sequence[0:10]}...' composed of {Ncodons} codons = {mRNA_seq_length} nucleotides ...")
        print(f"\t... has ~10^{round(total_combinations_EXPONENTIAL_PART_BASE_10 , 2)} synonymous mRNA sequences associated.")
        print(f"\tAverage degeneration degree per codon = {round(AverageDegenerationDegreePerCodon , 2)}")
        print(f"\nNucleotide A occurs {A_NucleotideCount} times = {A_NucleotidePercentage}%.")
        print(  f"Nucleotide C occurs {C_NucleotideCount} times = {C_NucleotidePercentage}%.")
        print(  f"Nucleotide G occurs {G_NucleotideCount} times = {G_NucleotidePercentage}%.")
        print(  f"Nucleotide T occurs {T_NucleotideCount} times = {T_NucleotidePercentage}%.")
        print(  f"Nucleotide U occurs {U_NucleotideCount} times = {U_NucleotidePercentage}%.")
        print("\t------------------")
        
        # Printing the information about the codon occurrences
        print(f"\nTable of occurrences and % of apparition of every codon in this sequence: (sorted by '{ORDER_TYPE}')")
        print("\t------------------")
        print("\tName: #     ; %")
        if   ORDER_TYPE == 'SORT_by_NAME': List_CodonCounts_SORTED = sorted(Dict_CodonCounts.items(), key=lambda x: x[0]) # Sort by codon name (key)
        elif ORDER_TYPE == 'SORT_by_%'   : List_CodonCounts_SORTED = sorted(Dict_CodonCounts.items(), key=lambda x: x[1][1], reverse=True) # Sort by % (value[1])
        for codon, (TotalCount , OccurrencePercentage , CodonPositions) in List_CodonCounts_SORTED: print(f"\t{codon} : {TotalCount} \t; {round(OccurrencePercentage, 2)}")
        print("\tName: #     ; %")
        print("\t------------------")
    
    # Ploting each codon occurrence
    if ShowCombinationEvolutionPlot == True:
        plt.figure(figsize = (DefaultFigureWidth , DefaultFigureHeight) , dpi = DefaultDPIResolution , facecolor = DefaultFaceColor , edgecolor = DefaultEdgeColor , layout = DefaultLayout)
        plt.title("Distribution")
        plt.legend()
        plt.grid(visible = DefaultGridValue , which = 'both' , alpha = DefaultGridOpacity , color = DefaultGridColor , linestyle = DefaultGridLineStyle , linewidth = DefaultGridLineWidth)
        plt.show()
    
    # Returns the results
    return mRNA_seq_FILTERED , mRNA_seq_length , List_CodonCounts_SORTED , total_combinations , total_combinations_EXPONENTIAL_PART_BASE_10 , AverageDegenerationDegreePerCodon

def read_csv(filename , delimiter , STARTING_LINE_FORCED = 0, STOPPING_LINE_FORCED = -1, PercentageStep = 10):
    # Open the Excel file
    try:
        with open(filename, 'r', newline='') as csvfile :
            csv_reader = csv.reader(csvfile, delimiter = delimiter)
            
            # Initializing the array containing the data
            Array = np.zeros((len(csv_reader) , len(csv_reader[0])))
        
            # Loop through the rows
            for csv_line_number_tmp , row in enumerate(csv_reader):
                print(row)
                
                # Loop through the columns
                for csv_column_number_tmp in range(len(row[csv_line_number_tmp])):
                    Cell_i_j = row[csv_line_number_tmp][csv_column_number_tmp]
                    Array[csv_line_number_tmp][csv_column_number_tmp] = Cell_i_j
                    print(f"Cell_i_j = '{Cell_i_j}'")
    
    except Exception as error_message : print(f"Error reading the .xlsx file ('read_csv' function): {error_message}")
    print("\nCSV READING DONE !")
    return Array

def read_xlsx(file_path , MUTE_MESSAGES = True):
    # Open the Excel file
    try:
        workbook = openpyxl.load_workbook(file_path)
        
        # Counting the number of sheets, maximum number of rows per sheet, maximum number ocolumns per sheet
        Nb_excel_sheets       = len(workbook.sheetnames)
        Nb_rows_MAX     = 1
        Nb_columns_MAX  = 1
        for sheet_number , sheet_name   in enumerate(workbook.sheetnames): # Loop through each sheet in the workbook
            for row_number , row        in enumerate(workbook[sheet_name]): # Loop through each row in the current sheet
                if Nb_rows_MAX < row_number + 1 : Nb_rows_MAX = row_number + 1
                for column_number , column in enumerate(row): # Loop through each column in the current row
                    if Nb_columns_MAX < column_number + 1 : Nb_columns_MAX = column_number + 1
                    
        # Initializing the array containing the data
        Array = np.empty((Nb_excel_sheets , Nb_rows_MAX , Nb_columns_MAX) , dtype = object)
        
        # Loop through each sheet in the workbook
        for sheet_number , sheet_name   in enumerate(workbook.sheetnames): # Loop through each sheet in the workbook
            for row_number , row        in enumerate(workbook[sheet_name]): # Loop through each row in the current sheet
                for column_number , column in enumerate(row): # Loop through each column in the current row
                    
                    # Reading the cell value and adding it to the resulting array
                    Cell_i_j = workbook[sheet_name].cell(row = row_number+1 , column = column_number+1).value
                    
                    # Registering the value of the cell into the array
                    Array[sheet_number][row_number][column_number] = Cell_i_j
                    
                    # Eventually printing the registered cell
                    if MUTE_MESSAGES == False : print(f"\t\tCell ({row_number} ; {column_number}) = '{Cell_i_j}'")
    
    except Exception as error_message : print(f"Error reading the .xlsx file ('read_xlsx' function):\n\t'{error_message}'")
    print(f"\n\tReading CSV file done : '{file_path}'")
    return Array

def clean_empty_rows_and_columns(Array):
    
    NOT_EMPTY_DATA = [[] for u in range(len(Array))]
    
    # Loop through each sheet in the array - eliminating the empty rows first
    for sheet_number, sheet in enumerate(Array):
        for row_number, row in enumerate(sheet): # Loop through each row in the current sheet
            # Check if the row contains at least one non-empty value
            if not np.all([elem in [None, np.nan, ''] for elem in row]) : NOT_EMPTY_DATA[sheet_number].append(row)

    # Loop through each sheet in the array - eliminating the empty columns
    for sheet_number, sheet in enumerate(Array):
        Nb_columns_already_deleted = 0
        for column_number in range(sheet.shape[1]): # Loop through each column in the current sheet
            # column = sheet[:, column_number]
            column = [row[column_number] for row in sheet]

            # Check if the column is empty this time ...
            if np.all([elem in [None, np.nan, ''] for elem in column]) :
                
                # Delete the column by removing the column_number from each row
                for row_number , row in enumerate(NOT_EMPTY_DATA[sheet_number]) :
                    NOT_EMPTY_DATA[sheet_number][row_number] = list(NOT_EMPTY_DATA[sheet_number][row_number])
                    NOT_EMPTY_DATA[sheet_number][row_number].pop(column_number - Nb_columns_already_deleted)
                
                # Incrementing the number of columns already deleted for this sheet only
                Nb_columns_already_deleted += 1 
                    
    return np.array(NOT_EMPTY_DATA)

def extracting_time_and_concentration(NOT_EMPTY_DATA,
                                      expected_data_Nb_per_excel_sheet              = 3,
                                      expected_row_for_datalabels                   = 0,
                                      expected_row_for_timelabels                   = 1,
                                      expected_row_for_mRNA_dosages                 = 1,
                                      expected_row_for_time_START                   = 2,
                                      expected_row_for_protein_concentration_START  = 2,
                                      expected_column_offset_for_datalabels         = 0,
                                      expected_column_offset_for_timelabels         = 0,
                                      expected_column_offset_for_mRNA_dosages       = 1,
                                      expected_columns_per_data                     = 2,
                                      BOOL_ReplaceNonBreakingSpace_by_space         = True):
    
    # Initializing the total number of curves (named '...total_data' but equivalent)
    Nb_excel_sheets     = len(NOT_EMPTY_DATA)
    Nb_total_curves       = Nb_excel_sheets * expected_data_Nb_per_excel_sheet
    # Nb_total_columns    = expected_data_Nb_per_excel_sheet * expected_columns_per_data
    
    # Initializing the arrays furtherely containing the different labels, texts, datas
    List_of_all_datalabels              = np.empty(shape = Nb_total_curves , dtype = object)
    List_of_all_timelabels              = np.empty(shape = Nb_total_curves , dtype = object)
    List_of_all_mRNA_dosages            = np.empty(shape = Nb_total_curves , dtype = object)
    List_of_all_time                    = np.empty(shape = Nb_total_curves , dtype = object)
    List_of_all_protein_concentrations  = np.empty(shape = Nb_total_curves , dtype = object)
    
    # Initializing the different counts to save and register the labels, strings, texts, datas successively
    cpt_datalabels      = 0
    cpt_timelabels      = 0
    cpt_mRNA_dosages    = 0
    cpt_time_points     = 0
    cpt_protein_concs   = 0
    
    # Loop through each sheet in the array - eliminating the empty rows first
    for sheet_number, sheet in enumerate(NOT_EMPTY_DATA):
        for row_number, row in enumerate(sheet): # Loop through each row in the current sheet
            for column_number in range(sheet.shape[1]): # Loop through each column in the current sheet
                
                # Registering the data labels (as text)
                if row_number == expected_row_for_datalabels and ((column_number-expected_column_offset_for_datalabels)/expected_columns_per_data == (column_number-expected_column_offset_for_datalabels)//expected_columns_per_data):
                    
                    # Eventually replacing the eventual non-breaking space caracters by a space (non-breaking space caracter occuring due to the use of openpxl library apparently ...)
                    Cell_i_j = row[column_number]
                    if isinstance(Cell_i_j , str) and BOOL_ReplaceNonBreakingSpace_by_space == True: Cell_i_j = Cell_i_j.replace('\xa0', ' ')
                    List_of_all_datalabels[cpt_datalabels] = Cell_i_j
                    cpt_datalabels = cpt_datalabels + 1
                
                # Registering the time labels (as text)
                if row_number == expected_row_for_timelabels and ((column_number-expected_column_offset_for_timelabels)/expected_columns_per_data == (column_number-expected_column_offset_for_timelabels)//expected_columns_per_data):
                    
                    # Eventually replacing the eventual non-breaking space caracters by a space (non-breaking space caracter occuring due to the use of openpxl library apparently ...)
                    Cell_i_j = row[column_number]
                    if isinstance(Cell_i_j , str) and BOOL_ReplaceNonBreakingSpace_by_space == True: Cell_i_j = Cell_i_j.replace('\xa0', ' ')
                    List_of_all_timelabels[cpt_timelabels] = Cell_i_j
                    cpt_timelabels = cpt_timelabels + 1
                
                # Registering the mRNA doses (as text)
                if row_number == expected_row_for_mRNA_dosages and ((column_number-expected_column_offset_for_mRNA_dosages)/expected_columns_per_data == (column_number-expected_column_offset_for_mRNA_dosages)//expected_columns_per_data):
                    
                    # Eventually replacing the eventual non-breaking space caracters by a space (non-breaking space caracter occuring due to the use of openpxl library apparently ...)
                    Cell_i_j = row[column_number]
                    if isinstance(Cell_i_j , str) and BOOL_ReplaceNonBreakingSpace_by_space == True: Cell_i_j = Cell_i_j.replace('\xa0', ' ')
                    List_of_all_mRNA_dosages[cpt_mRNA_dosages] = Cell_i_j
                    cpt_mRNA_dosages = cpt_mRNA_dosages + 1
                
                # Registering the time points
                if row_number == expected_row_for_time_START and ((column_number-expected_column_offset_for_timelabels)/expected_columns_per_data == (column_number-expected_column_offset_for_timelabels)//expected_columns_per_data):
                        
                    List_of_all_time[cpt_time_points] = np.array([sheet[u][column_number] for u in range(expected_row_for_time_START,len(sheet))])
                    cpt_time_points = cpt_time_points + 1
                
                # Registering the protein concentrations in time
                if row_number == expected_row_for_protein_concentration_START and ((column_number-expected_column_offset_for_mRNA_dosages)/expected_columns_per_data == (column_number-expected_column_offset_for_mRNA_dosages)//expected_columns_per_data):
                        
                    List_of_all_protein_concentrations[cpt_protein_concs] = np.array([sheet[u][column_number] for u in range(expected_row_for_protein_concentration_START,len(sheet))])
                    cpt_protein_concs = cpt_protein_concs + 1
    
    return List_of_all_datalabels , List_of_all_timelabels , List_of_all_mRNA_dosages , List_of_all_time , List_of_all_protein_concentrations 

# # # # # # # # # # # # # # # # # # # # # # # # # # # #
def normalize_array(Array , DEF_MIN = 0 , DEF_MAX = 1):
    
    if Array[0].shape == (): # This means the array is composed of only one 1D array
        # Find the minimum and maximum of the array
        arr_min = np.min(Array)
        arr_max = np.max(Array)
        
    elif Array[0].shape >= (1,): # This means the array is composed of multiple 1D subarrays
        arr_min = Array[0][0]
        arr_max = arr_min
        for sub_array_num , sub_array in enumerate(Array):
            if np.min(sub_array) < arr_min : arr_min = np.min(sub_array)
            if arr_max < np.max(sub_array) : arr_max = np.max(sub_array)
    
    else : # Too many dimensions => update this function.
        raise(ValueError("FUNCTION 'normalize_array' : # Too many dimensions => update this function to properly work ..."))
    
    # Avoid division by zero if the array has a constant value
    if arr_min == arr_max : return np.full(Array.shape , DEF_MIN) # or DEF_MAX; choose one
    
    # Normalize the array to the range [DEF_MIN, DEF_MAX]
    normalized_array = DEF_MIN + (Array - arr_min) * (DEF_MAX - DEF_MIN) / (arr_max - arr_min)
    return normalized_array

def Generate_mRNA_t_BEFORE_INJECTION(alpha_mRNA, beta_mRNA, t, t_injection):
    # The mRNA model is following this physico-chemical equation
    #   d([mRNA](t))/dt = alpha_mRNA - beta_mRNA * [mRNA](t) # WHERE alpha_mRNA and beta_mRNA are > 0 (not equal to 0)
    
    mRNA_t = alpha_mRNA/beta_mRNA * (1 - np.exp(-beta_mRNA * (t - t_injection)))
    return max(0. , mRNA_t) # always returns a psitive concentration

def Generate_mRNA_t_AFTER_INJECTION(alpha_mRNA, beta_mRNA, mRNA_inj, t, tinj):
    # The mRNA model is following this physico-chemical equation
    #   d([mRNA](t))/dt = alpha_mRNA - beta_mRNA * [mRNA](t) # WHERE alpha_mRNA and beta_mRNA are > 0 (not equal to 0)
    
    mRNA_t = alpha_mRNA/beta_mRNA + mRNA_inj * np.exp(-beta_mRNA * (t - tinj))
    return max(0. , mRNA_t) # always returns a psitive concentration

def Generate_List_of_mRNA_t_BEFORE_INJECTION(alpha_mRNA, beta_mRNA, mRNA_t_injection, List_of_time, t_injection):
    # The mRNA model is following this physico-chemical equation
    #   d([mRNA](t))/dt = alpha_mRNA - beta_mRNA * [mRNA](t) # WHERE alpha_mRNA and beta_mRNA are > 0 (not equal to 0)
    
    # Initializing the resulting list of mRNA concentrations
    List_of_mRNA_concentrations = np.zeros(len(List_of_time))
    for i , time_t in enumerate(List_of_time) : List_of_mRNA_concentrations[i] = Generate_mRNA_t_BEFORE_INJECTION(alpha_mRNA, beta_mRNA, time_t, t_injection)
    return List_of_mRNA_concentrations

def Generate_List_of_mRNA_t_AFTER_INJECTION(alpha_mRNA, beta_mRNA, mRNA_inj, List_of_time, tinj):
    # The mRNA model is following this physico-chemical equation
    #   d([mRNA](t))/dt = alpha_mRNA - beta_mRNA * [mRNA](t) # WHERE alpha_mRNA and beta_mRNA are > 0 (not equal to 0)
    
    # Initializing the resulting list of mRNA concentrations
    List_of_mRNA_concentrations = np.zeros(len(List_of_time))
    for i , time_t in enumerate(List_of_time) : List_of_mRNA_concentrations[i] = Generate_mRNA_t_AFTER_INJECTION(alpha_mRNA, beta_mRNA, mRNA_inj, time_t, tinj)
    return List_of_mRNA_concentrations

def Generate_protein_t_BEFORE_INJECTION(alpha_protein, beta_protein, t, t_injection, alpha_mRNA, beta_mRNA):
    # The mRNA model is following this physico-chemical equation
    #   d([mRNA](t))/dt = alpha_mRNA - beta_mRNA * [mRNA](t) # WHERE alpha_mRNA and beta_mRNA are > 0 (not equal to 0)
    
    # The protein model is following this physico-chemical equation
    #   d([protein](t))/dt = alpha_protein * [mRNA](t) - beta_protein * [protein](t) # WHERE alpha_protein and beta_protein are > 0 and alpha_mRNA and beta_mRNA are > 0 (not equal to 0)
    
    delta_beta  = beta_protein - beta_mRNA
    mRNA_eq     = alpha_mRNA / beta_mRNA
    protein_eq  = mRNA_eq * alpha_protein / beta_protein # Protein concentration at equilibrium
    protein_t   = protein_eq + (alpha_protein * mRNA_eq / delta_beta - protein_eq) * np.exp(-beta_protein * (t - t_injection)) - alpha_protein * mRNA_eq / delta_beta * np.exp(-beta_mRNA * (t - t_injection))
    return max(0. , protein_t) # always returns a positive concentration

def Generate_protein_t_AFTER_INJECTION(alpha_protein, beta_protein, t, tinj, alpha_mRNA, beta_mRNA , mRNA_inj):
    # The mRNA model is following this physico-chemical equation
    #   d([mRNA](t))/dt = alpha_mRNA - beta_mRNA * [mRNA](t) # WHERE alpha_mRNA and beta_mRNA are > 0 (not equal to 0)
    
    # The protein model is following this physico-chemical equation
    #   d([protein](t))/dt = alpha_protein * [mRNA](t) - beta_protein * [protein](t) # WHERE alpha_protein and beta_protein are > 0 and alpha_mRNA and beta_mRNA are > 0 (not equal to 0)
    
    delta_beta  = beta_protein - beta_mRNA
    mRNA_eq     = alpha_mRNA / beta_mRNA
    protein_eq  = mRNA_eq * alpha_protein / beta_protein # Protein concentration at equilibrium
    protein_t   = protein_eq - alpha_protein * mRNA_inj / delta_beta * np.exp(-beta_protein * (t - tinj)) + alpha_protein * mRNA_inj / delta_beta * np.exp(-beta_mRNA * (t - tinj))
    return max(0. , protein_t) # always returns a positive concentration

def Generate_List_of_protein_t_BEFORE_INJECTION(alpha_protein, beta_protein, List_of_time, t_injection, alpha_mRNA, beta_mRNA):
    # The mRNA model is following this physico-chemical equation
    #   d([mRNA](t))/dt = alpha_mRNA - beta_mRNA * [mRNA](t) # WHERE alpha_mRNA and beta_mRNA are > 0 (not equal to 0)
    
    # The protein model is following this physico-chemical equation
    #   d([protein](t))/dt = alpha_protein * [mRNA](t) - beta_protein * [protein](t) # WHERE alpha_protein and beta_protein are > 0 and alpha_mRNA and beta_mRNA are > 0 (not equal to 0)
    
    # Initializing the resulting list of mRNA concentrations
    List_of_protein_concentrations = np.zeros(len(List_of_time))
    for i , time_t in enumerate(List_of_time) : List_of_protein_concentrations[i] = Generate_protein_t_BEFORE_INJECTION(alpha_protein, beta_protein, time_t, t_injection, alpha_mRNA, beta_mRNA)
    return List_of_protein_concentrations

def Generate_List_of_protein_t_AFTER_INJECTION(alpha_protein, beta_protein, List_of_time, tinj, alpha_mRNA, beta_mRNA, mRNA_inj):
    # The mRNA model is following this physico-chemical equation
    #   d([mRNA](t))/dt = alpha_mRNA - beta_mRNA * [mRNA](t) # WHERE alpha_mRNA and beta_mRNA are > 0 (not equal to 0)
    
    # The protein model is following this physico-chemical equation
    #   d([protein](t))/dt = alpha_protein * [mRNA](t) - beta_protein * [protein](t) # WHERE alpha_protein and beta_protein are > 0 and alpha_mRNA and beta_mRNA are > 0 (not equal to 0)
    
    # Initializing the resulting list of mRNA concentrations
    List_of_protein_concentrations = np.zeros(len(List_of_time))
    for i , time_t in enumerate(List_of_time) : List_of_protein_concentrations[i] = Generate_protein_t_AFTER_INJECTION(alpha_protein, beta_protein, time_t, tinj, alpha_mRNA, beta_mRNA , mRNA_inj)
    return List_of_protein_concentrations

def calculate_protein_theoretical_inflexion_timepoint(tinj, beta_protein, beta_mRNA): return tinj + np.log(beta_protein / beta_mRNA) / (beta_protein - beta_mRNA)

def calculate_theoretical_maximum_concentration(alpha_protein, beta_protein, alpha_mRNA, beta_mRNA, mRNA_inj):
    
    Peq         = alpha_protein * alpha_mRNA / (beta_protein * beta_mRNA)
    coeff       = alpha_protein * mRNA_inj / (beta_protein - beta_mRNA)
    delta_betas = beta_protein - beta_mRNA
    ratio_betas = beta_protein / beta_mRNA
    power_1     = - beta_mRNA    * delta_betas
    power_2     = - beta_protein * delta_betas
    term_1      = pow(ratio_betas , power_1)
    term_2      = pow(ratio_betas , power_2)
    
    protein_concentration_THEOR_MAX = Peq + coeff * (term_1 - term_2)
    return protein_concentration_THEOR_MAX

def find_Tau_half_life_mRNA_from_optimal_parameters(t_injection, alpha_mRNA_OPTI, beta_mRNA_OPTI, mRNA_t_injection):
    
    # Compute key concentrations
    mRNA_at_t_equilibrium               = alpha_mRNA_OPTI / beta_mRNA_OPTI
    mRNA_at_Tau_half_life_mRNA          = mRNA_at_t_equilibrium + (mRNA_t_injection / 2)
    Tau_half_life_mRNA                  = - math.log(1/2 - mRNA_at_t_equilibrium / mRNA_t_injection) / beta_mRNA_OPTI
    t_injection_plus_Tau_half_life_mRNA = t_injection + Tau_half_life_mRNA
    
    return t_injection_plus_Tau_half_life_mRNA, Tau_half_life_mRNA, mRNA_at_Tau_half_life_mRNA

def find_Tau_half_life_protein_from_optimal_parameters(alpha_protein_OPTI, beta_protein_OPTI, t_injection, alpha_mRNA_OPTI, beta_mRNA_OPTI, mRNA_t_injection, Number_of_iterations_MAX, MAXIMUM_TIMEPOINT, precision_required_as_decimal_value):
    
    # Compute key concentrations
    P_at_t_equilibrium = alpha_protein_OPTI * alpha_mRNA_OPTI / (beta_protein_OPTI * beta_mRNA_OPTI)
    t_inflexion        = calculate_protein_theoretical_inflexion_timepoint(t_injection, beta_protein_OPTI, beta_mRNA_OPTI)
    P_at_t_inflexion   = Generate_protein_t_AFTER_INJECTION(alpha_protein_OPTI, beta_protein_OPTI, t_inflexion, t_injection, alpha_mRNA_OPTI, beta_mRNA_OPTI , mRNA_t_injection)
    P_at_Tau_half_life_protein = (P_at_t_equilibrium + P_at_t_inflexion) / 2
    
    protein_concentration_at_last_timepoint = Generate_protein_t_AFTER_INJECTION(alpha_protein_OPTI, beta_protein_OPTI, MAXIMUM_TIMEPOINT, t_injection, alpha_mRNA_OPTI, beta_mRNA_OPTI , mRNA_t_injection)
    if protein_concentration_at_last_timepoint > P_at_Tau_half_life_protein:
        raise ValueError(f"The last timepoint set for searching the Tau half-life in the associated range is too low...\n\tThe protein concentration at this last point is still too high compared to the protein concentration expected at half-life: {round(protein_concentration_at_last_timepoint,2)} (at last timepoint) > {round(P_at_Tau_half_life_protein,2)} (at tau half-life)")
    else:
        
        # Dichotomy search
        time_lower_bound = t_inflexion
        time_upper_bound = MAXIMUM_TIMEPOINT
        tolerance = pow(10 , -precision_required_as_decimal_value)
        
        for _ in range(Number_of_iterations_MAX):
            t_middle = (time_lower_bound + time_upper_bound) / 2
            protein_at_t_middle = Generate_protein_t_AFTER_INJECTION(alpha_protein_OPTI, beta_protein_OPTI, t_middle, t_injection, alpha_mRNA_OPTI, beta_mRNA_OPTI , mRNA_t_injection)
            if abs(protein_at_t_middle - P_at_Tau_half_life_protein) <= tolerance:
                return t_inflexion, P_at_t_inflexion, round(t_middle, precision_required_as_decimal_value), P_at_Tau_half_life_protein
            if protein_at_t_middle                           > P_at_Tau_half_life_protein:  time_lower_bound = t_middle
            elif P_at_Tau_half_life_protein                          > protein_at_t_middle: time_upper_bound = t_middle
        
        # If max iterations reached, return the best guess
        Tau_half_life_protein_approx = round((time_lower_bound + time_upper_bound) / 2, precision_required_as_decimal_value)
        Warning(f"NOT ENOUGH ITERATIONS ({Number_of_iterations_MAX}) to find Tau half-life with the precision needed ({precision_required_as_decimal_value} decimal integers).\n\tTau half-life is around {Tau_half_life_protein_approx} seconds")
        return t_inflexion, P_at_t_inflexion, Tau_half_life_protein_approx, P_at_Tau_half_life_protein

def calculate_MSE(THEOR_protein_concentration , EXPERIMENTAL_protein_concentration):
    ## EXAMPLE : dMSE_dalpha_protein = d_MSE_over_d_alpha_protein(EXPERIMENTAL_protein_concentration , alpha_protein, beta_protein, protein_t_injection, List_of_time, t_injection, alpha_mRNA, beta_mRNA, mRNA_t_injection)
    # The mRNA model is following this physico-chemical equation
    #   d([mRNA](t))/dt = alpha_mRNA - beta_mRNA * [mRNA](t) # WHERE alpha_mRNA and beta_mRNA are > 0 (not equal to 0)
    # The protein model is following this physico-chemical equation
    #   d([protein](t))/dt = alpha_protein * [mRNA](t) - beta_protein * [protein](t) # WHERE alpha_protein and beta_protein are > 0 and alpha_mRNA and beta_mRNA are > 0 (not equal to 0)
    
    # For loop to determine all the derivatives of the MSE over alpha_protein
    MSE = 0
    for i , EXP_protein_concentration in enumerate(EXPERIMENTAL_protein_concentration):
        P_ti_THEOR  = THEOR_protein_concentration[i]
        P_ti_EXP    = EXPERIMENTAL_protein_concentration[i]
        MSE         = MSE + pow(P_ti_THEOR - P_ti_EXP , 2)
    return MSE

def mrna_molar_mass(sequence):
    """ Computes the total molar mass of an mRNA sequence.
    Parameters : sequence (str): The mRNA sequence (a string containing A, U, G, C).
    Returns : float: The total molar mass of the mRNA sequence in g/mol (Daltons).
    """
    # Molar masses of individual nucleotides (in g/mol or Daltons)
    molar_masses = {'A': 329.2, # Adenine
                    'U': 306.2, # Uracil
                    'G': 345.2, # Guanine
                    'C': 305.2} # Cytosine
    
    # Iterate through the sequence and sum up the molar masses
    total_mass = 0.0
    for nucleotide in sequence:
        if nucleotide in molar_masses : total_mass += molar_masses[nucleotide]
        else : raise ValueError(f"Invalid nucleotide '{nucleotide}' in sequence.")
    return total_mass

def find_Tau_life_longing_at_p_percent_in_EXP_DATA(EXPERIMENTAL_protein_concentration , List_of_time , p_percentage):
    max_index = np.argmax(EXPERIMENTAL_protein_concentration)  # Index of the highest value
    max_value = EXPERIMENTAL_protein_concentration[max_index]  # Highest value itself
    threshold = p_percentage / 100 * max_value  # Compute the threshold
    
    # Find the first index after max_index where the value drops below the threshold
    for i in range(max_index + 1, len(EXPERIMENTAL_protein_concentration)):
        if EXPERIMENTAL_protein_concentration[i] <= threshold : return List_of_time[i] - List_of_time[max_index] # Return the delay (time difference)
    
    return None # Return None if no such moment is found

# Returns both the normalized value and exponent as strings.
def extract_value_and_exponent(value):
    # Example of return: integer_part , exponent_part = extract_value_and_exponent(value)
    
    if value == 0: return ["0", "0"] # Special case for zero
    if np.isnan(value): return ["-1", "0"] # Special case for NaN value
    exponent = math.floor(math.log10(abs(value))) # Find power of 10
    normalized_value = value / (10 ** exponent) # Normalize to [1, 10)
    return str(normalized_value.item()) , str(exponent) # Return as strings

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
def save_plot_with_timestamp(directory = r"\." , figure_name = "figure_saved" , filename_prefix = "plot" , timestamp = datetime.now().strftime("%Y-%m-%d-%H-%M-%S") , Print_message = True):
  if not os.path.exists(directory):
      try: os.makedirs(directory)
      except:
          if Print_message: print(f"func save_plot_with_timestamp: ERROR: an error occured while trying to create the saving directory as '{directory}'")
          return False
  else:
      try:
          filename  = f"{filename_prefix}_{timestamp}.png"
          filepath  = os.path.join(directory , filename)
          figure_name.savefig(filepath)
          if Print_message: print(f"\nPlot saved as : {filepath}")
          return True
      except:
          if Print_message: print(f"func save_plot_with_timestamp: ERROR: an error occured while trying to save the figure as '{filepath}'")
          return False

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
def Show_Progression(TIME_START_LOOP                       = 0, # time start in seconds
                     TIME_NOW                              = 1, # current time in seconds
                     Number_of_tabulations_before_sentence = 1,
                     DescriptiveWord                       = 'COMPUTATION',
                     current_index                         = 0,
                     index_start                           = 0,
                     index_stop                            = 0,
                     PreviousReadingPercentage             = 0,
                     PercentageStep                        = 10,
                     round_digit                           = 2,
                     BOOL_do_not_show_time_left            = False):
    
    if index_start == index_stop : return 100
    # step                 = 1 if index_stop > index_start else -1
    abs_delta_start_stop = max(index_stop , index_start) - min(index_stop , index_start) + 1
    if index_start < index_stop  and index_start <  0 : index_normalized = current_index - index_start
    if index_start < index_stop  and index_start >= 0 : index_normalized = current_index - index_start
    if index_start >= index_stop and index_start <  0 : index_normalized = index_start - current_index
    if index_start >= index_stop and index_start >= 0 : index_normalized = abs_delta_start_stop - (current_index - index_stop)
    
    # Calculating the percentage of steps done compared to the total to do
    ReadingPercentage = round(100 * index_normalized / abs_delta_start_stop, round_digit)
    
    # Only printing the message when the current percentage has been increased more than PercentageStep compared to the PreviousReadingPercentage
    if PreviousReadingPercentage + PercentageStep <= ReadingPercentage :
        
        # Calculating the remaining computational time
        elapsed_time_tmp = time.time() - TIME_START_LOOP # number of elapsed MINUTES
        if ReadingPercentage == 0 or BOOL_do_not_show_time_left: Message_time_remaining_STR = ''
        else:
            remaining_time_tmp = (100 - ReadingPercentage) * elapsed_time_tmp / ReadingPercentage
            remaining_time_DD_tmp = remaining_time_tmp                    // (60 * 60 * 24)
            remaining_time_HH_tmp = (remaining_time_tmp % (60 * 60 * 24)) // (60 * 60)
            remaining_time_MN_tmp = (remaining_time_tmp % (60 * 60))      // 60
            remaining_time_SS_tmp = round(remaining_time_tmp % 60)
            
            string_for_DD = '' if remaining_time_DD_tmp == 0 else f'{remaining_time_DD_tmp}day '
            string_for_HH = '' if remaining_time_HH_tmp == 0 else f'{remaining_time_HH_tmp}h '
            string_for_MN = '' if remaining_time_MN_tmp == 0 else f'{remaining_time_MN_tmp}mn '
            string_for_SS = '' if remaining_time_SS_tmp == 0 else f'{remaining_time_SS_tmp}sec'
            Message_time_remaining_STR = f" ; Remaining time = {string_for_DD}{string_for_HH}{string_for_MN}{string_for_SS}"
            if Message_time_remaining_STR == ' ; Remaining time = ': Message_time_remaining_STR = ''
        
        # Constructs the final message
        initial_tabulations       = '\t' * Number_of_tabulations_before_sentence
        PreviousReadingPercentage = round(ReadingPercentage , round_digit)
        print(f"{initial_tabulations}{DescriptiveWord} : {index_normalized}/{abs_delta_start_stop} = {PreviousReadingPercentage}%" + Message_time_remaining_STR)
    
    return PreviousReadingPercentage

def PrintElapsedTime(start_time, end_time , MUTE = False):
    # Conversion constants in seconds
    units = {
        "Y" : 365.25 * 24 * 60 * 60,
        "M" : (365.25 / 12) * 24 * 60 * 60,
        "d" : 24 * 60 * 60,
        "h" : 60 * 60,
        "mn": 60,
        "s" : 1,
        "ms": 1e-3,
        "μs": 1e-6,
        "ns": 1e-9}
    
    elapsed_time    = end_time - start_time
    remaining_time  = elapsed_time
    results         = {}

    # Calculate each time component
    for unit , factor in units.items() :
        results[unit]   = int(remaining_time // factor)
        remaining_time %= factor

    # Format the elapsed time for printing
    if   elapsed_time <= 1e-6       : formatted_time = f"{results['ns']}ns"
    elif elapsed_time <= 1e-3       : formatted_time = f"{results['μs']}μs {results['ns']}ns"
    elif elapsed_time <= 1          : formatted_time = f"{results['ms']}ms {results['μs']}μs {results['ns']}ns"
    elif elapsed_time <= 60         : formatted_time = f"{results['s']}s {results['ms']}ms {results['μs']}μs {results['ns']}ns"
    elif elapsed_time <= 3600       : formatted_time = f"{results['mn']}mn {results['s']}s {results['ms']}ms {results['μs']}μs {results['ns']}ns"
    elif elapsed_time <= 86400      : formatted_time = f"{results['h']}h {results['mn']}mn {results['s']}s {results['ms']}ms {results['μs']}μs {results['ns']}ns"
    elif elapsed_time <= units["M"] : formatted_time = f"{results['d']}d {results['h']}h {results['mn']}mn {results['s']}s {results['ms']}ms {results['μs']}μs {results['ns']}ns"
    elif elapsed_time <= units["Y"] : formatted_time = f"{results['M']}M {results['d']}d {results['h']}h {results['mn']}mn {results['s']}s {results['ms']}ms {results['μs']}μs {results['ns']}ns"
    else : formatted_time = f"{results['Y']}Y {results['M']}M {results['d']}d {results['h']}h {results['mn']}mn {results['s']}s {results['ms']}ms {results['μs']}μs {results['ns']}ns"
    
    if not MUTE : print(f"\nTotal elapsed time = {formatted_time}")

    # Return the elapsed time components
    return formatted_time , (elapsed_time, *[results[unit] for unit in ["Y", "M", "d", "h", "mn", "s", "ms", "μs", "ns"]])

def ReturnRGBColorCode(color_string = 'red') :
  if    color_string == 'white'   or color_string == 'blanc'  : return [1   , 1   , 1]
  elif  color_string == 'gray'    or color_string == 'gris'   : return [0.5 , 0.5 , 0.5]
  elif  color_string == 'black'   or color_string == 'noir'   : return [0   , 0   , 0]
  elif  color_string == 'red'     or color_string == 'rouge'  : return [1   , 0   , 0]
  elif  color_string == 'orange'                              : return [1   , 0.5 , 0]
  elif  color_string == 'yellow'  or color_string == 'jaune'  : return [1   , 1   , 0]
  elif  color_string == 'apple'   or color_string == 'pomme'  : return [0.5 , 1   , 0]
  elif  color_string == 'green'   or color_string == 'vert'   : return [0   , 1   , 0]
  elif  color_string == 'cyan'                                : return [0   , 1   , 1]
  elif  color_string == 'blue'    or color_string == 'bleu'   : return [0   , 0   , 1]
  elif  color_string == 'indigo'                              : return [0.5 , 0   , 1]
  elif  color_string == 'purple'  or color_string == 'violet' : return [1   , 0   , 1]
  elif  color_string == 'magenta'                             : return [1   , 0   , 0.5]
  else                                                        :
    print("WARNING in return 'ReturnRGBColorCode' function :\n {} is not recognized, Black is returned by default as [0 , 0 , 0] RGB color code.".format(color_string))
    return [0 , 0 , 0]

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
# # # # # # # REMOVING PLOTS BEFORE EXECUTION # # # # # # # # # # # # # # # # #
# Force-close all figures
plt.close('all')
TIME_START_PROGRAM = time.time() # Saving the starting time
print(f"Program Starting Time : {TIME_START_PROGRAM}\n")

###########################################################################################################
#### THE FOLLOWING BLOCK CAN BE COPY-PASTE IN EVERY CODE TO PLOT K colored curves composed of N VALUES ####
# Generate a RED TO BLACK list of colors (never white, the sequence is : rouge , orange , jaune , vert , cyan , bleu , indigo , violet , noir)
# Theoretically the number of triplet-color-values in Lcolors is equal to 5*(2^8-1) = 1275
Lcolors , index = [[]] , 0
for i     in range(1276)        : Lcolors.append([0,0,0])
for green in range(0,256,1)     : Lcolors[index] , index = [255 , green ,0]    , index + 1 # Color is RED # INCREASING THE VALUE of GREEN
for red   in range(254,-1,-1)   : Lcolors[index] , index = [red , 255   ,0]    , index + 1 # Color is YELLOW # DECREASING THE VALUE of RED
for blue  in range(1,256,1)     : Lcolors[index] , index = [0   , 255   ,blue] , index + 1 # Color is GREEN # INCREASING THE VALUE of BLUE
for green in range(254,-1,-1)   : Lcolors[index] , index = [0   , green ,255]  , index + 1 # Color is CYAN # DECREASING THE VALUE of GREEN
for blue  in range(254,0,-1)    : Lcolors[index] , index = [0   , 0     ,blue] , index + 1 # Color is BLUE # DECREASING THE VALUE of BLUE
for i     in range(len(Lcolors)-1,0,-1): # Color is BLACK # DELETING THE DOUBLE VALUES IN THE END
    if Lcolors[i] == Lcolors[i-1] : del Lcolors[i]
del Lcolors[-1]
for i in range(len(Lcolors)): # Then we normalize the values of Lcolors from [0:255] to [0:1]
    for j in range(len(Lcolors[0])): Lcolors[i][j]/=255
# Lcolors.reverse()
#### END OF THE GENERIC BLOCK ####
##################################

## Plots parameters
# FIGURE parameters
DefaultFigureWidth    = 6
DefaultFigureHeight   = DefaultFigureWidth * 9/16 # To automatically convert the plot size in a ratio of 16:9
DefaultDPIResolution  = 600 # 800 is enough for detailed plots on a PC screen
DefaultFaceColor      = ReturnRGBColorCode('white') # Spyder Black Background color is [25/100 , 35/100 , 45/100]
DefaultEdgeColor      = [0.1 , 0.1 , 0.1] # Spyder Black Background color is [25/100 , 35/100 , 45/100]
DefaultLayout         = 'tight' # 'constrained', 'compressed', 'tight'

# LINE / MARKER parameters
DefaultLineWidth  = 0.3
DefaultMarkerSize = 0.3
DefaultLineStyle  = '-'
DefaultFontSize   = 4

# GRID parameters
DefaultGridValue      = True # True # False
DefaultGridColor      = ReturnRGBColorCode('gray')
DefaultGridLineStyle  = '--'
DefaultGridLineWidth  = 0.4
DefaultGridOpacity    = 0.25

# TEXT parameters
SMALL_SIZE  = DefaultFontSize
MEDIUM_SIZE = SMALL_SIZE + 1
BIGGER_SIZE = MEDIUM_SIZE + 1

#####################################################################################
## INTERMEDIARY / NECESSARY / AUTOMATIC COMPUTATIONS BEFORE THE MAIN COMPUTATION PART
# Setting the different plot parameters before all plots
plt.rc('font'   , size      = SMALL_SIZE)       # controls default text sizes
plt.rc('axes'   , titlesize = SMALL_SIZE)       # fontsize of the axes title
plt.rc('axes'   , labelsize = SMALL_SIZE)       # fontsize of the x and y labels
plt.rc('xtick'  , labelsize = SMALL_SIZE)       # fontsize of the tick labels
plt.rc('ytick'  , labelsize = SMALL_SIZE)       # fontsize of the tick labels
plt.rc('legend' , fontsize  = SMALL_SIZE)       # legend fontsize
plt.rc('figure' , titlesize = BIGGER_SIZE)      # fontsize of the figure title
plt.rc('lines'  , linewidth = DefaultLineWidth)

##################
## MAIN PARAMETERS

# # # # # # # # # # # # #
## TXT reading parameters
PrintMessage                 = True
ShowCombinationEvolutionPlot = False
gene_sequence_FILE_FORMAT    = 'FASTA' # 'FASTA' # 'TEXT'
ProtSeq_StartingWord         = 'SEQUENCE=' # This string allows the function to recognize where does the protein sequence begins in case of FASTA format ... Otherwise for TEXT format it reads and considers all the string as the protein sequence by default ...
DNA_SEQ_FILENAME_TXT         = "Homo sapiens NPC intracellular cholesterol transporter 1 (NPC1), RefSeqGene on chromosome 18.txt"
DNA_GENE_SEQ_StartingWord    = 'RefSeqGene on chromosome 18'
ORDER_TYPE                   = 'SORT_by_%'  # Change to 'SORT_by_NAME' or 'SORT_by_%'

# # # # # # # # # # # # # #
# Getting the current directory
CurrentDirectory = os.getcwd()

# Setting the path leading to all data files
path_to_all_files = CurrentDirectory # path_to_all_files = r'C:\Users\...\Documents' # another path whether the user changes data location or not

## CSV reading parameters
Name_of_CSV_File = '20250121_Data analysis_HalfLife.xlsx'
filename         = os.path.join(path_to_all_files, Name_of_CSV_File)

# ALTERNATIVELY : if the searching range is discontinuous, one can directly put the 10th powers to work with in the following list ...
List_of_Power10_values_for_parameters = np.array([-15 , -13 , -11 , -9 , -7])
Nb_points_per_hour_THEORETICAL_CURVE  = 500 # Number of points per hour to plot for the optimized theoretical curve found after the optimization process

# Number of timepoints to generate for the tau half-life search
Number_of_iterations_MAX            = 100000
coef_t_limit_Tau_half_life_protein_finding  = 3 # This is the multiplicative coefficient in front of the experiment ending time point so we are sure to simulate the theoretical curve on a long-enough time duration to find Tau_half_life_protein (or Tau_p_percent more generally ...)
precision_required_as_decimal_value = 2 # precision for tau half-life in seconds (preicsion is in terms of seconds not in terms of hours)

#############################
## STARTING MAIN COMPUTATIONS
print("\n## STARTING MAIN COMPUTATIONS")

###################################################################
## OPENING AND REAGING THE CSV FILE AND THE GENE SEQUENCE TEXT FILE
print("## OPENING AND REAGING THE CSV FILE AND THE GENE SEQUENCE TEXT FILE")

# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
# NPC1 protein sequences (canonical = unvaried / original ; isoform = variated and genetically modified due to errors etc ...)
print("# Reading TXT file (mRNA / gene sequence reading)")
# NPC1_PROT_SEQ_canonical      = ">sp|O15118|NPC1_HUMAN NPC intracellular cholesterol transporter 1 OS=Homo sapiens OX=9606 GN=NPC1 PE=1 SV=2 SEQUENCE=MTARGLALGLLLLLLCPAQVFSQSCVWYGECGIAYGDKRYNCEYSGPPKPLPKDGYDLVQELCPGFFFGNVSLCCDVRQLQTLKDNLQLPLQFLSRCPSCFYNLLNLFCELTCSPRQSQFLNVTATEDYVDPVTNQTKTNVKELQYYVGQSFANAMYNACRDVEAPSSNDKALGLLCGKDADACNATNWIEYMFNKDNGQAPFTITPVFSDFPVHGMEPMNNATKGCDESVDEVTAPCSCQDCSIVCGPKPQPPPPPAPWTILGLDAMYVIMWITYMAFLLVFFGAFFAVWCYRKRYFVSEYTPIDSNIAFSVNASDKGEASCCDPVSAAFEGCLRRLFTRWGSFCVRNPGCVIFFSLVFITACSSGLVFVRVTTNPVDLWSAPSSQARLEKEYFDQHFGPFFRTEQLIIRAPLTDKHIYQPYPSGADVPFGPPLDIQILHQVLDLQIAIENITASYDNETVTLQDICLAPLSPYNTNCTILSVLNYFQNSHSVLDHKKGDDFFVYADYHTHFLYCVRAPASLNDTSLLHDPCLGTFGGPVFPWLVLGGYDDQNYNNATALVITFPVNNYYNDTEKLQRAQAWEKEFINFVKNYKNPNLTISFTAERSIEDELNRESDSDVFTVVISYAIMFLYISLALGHMKSCRRLLVDSKVSLGIAGILIVLSSVACSLGVFSYIGLPLTLIVIEVIPFLVLAVGVDNIFILVQAYQRDERLQGETLDQQLGRVLGEVAPSMFLSSFSETVAFFLGALSVMPAVHTFSLFAGLAVFIDFLLQITCFVSLLGLDIKRQEKNRLDIFCCVRGAEDGTSVQASESCLFRFFKNSYSPLLLKDWMRPIVIAIFVGVLSFSIAVLNKVDIGLDQSLSMPDDSYMVDYFKSISQYLHAGPPVYFVLEEGHDYTSSKGQNMVCGGMGCNNDSLVQQIFNAAQLDNYTRIGFAPSSWIDDYFDWVKPQSSCCRVDNITDQFCNASVVDPACVRCRPLTPEGKQRPQGGDFMRFLPMFLSDNPNPKCGKGGHAAYSSAVNILLGHGTRVGATYFMTYHTVLQTSADFIDALKKARLIASNVTETMGINGSAYRVFPYSVFYVFYEQYLTIIDDTIFNLGVSLGAIFLVTMVLLGCELWSAVIMCATIAMVLVNMFGVMWLWGISLNAVSLVNLVMSCGISVEFCSHITRAFTVSMKGSRVERAEEALAHMGSSVFSGITLTKFGGIVVLAFAKSQIFQIFYFRMYLAMVLLGATHGLIFLPVLLSYIGPSVNKAKSCATEERYKGTERERLLNF"
# NPC1_PROT_SEQ_isoform_2      = ">sp|O15118-2|NPC1_HUMAN Isoform 2 of NPC intracellular cholesterol transporter 1 OS=Homo sapiens OX=9606 GN=NPC1 SEQUENCE=MYVIMWITYMAFLLVFFGAFFAVWCYRKRYFVSEYTPIDSNIAFSVNASDKGTAWLLTSTFPSSPVLPGEASCCDPVSAAFEGCLRRLFTRWGSFCVRNPGCVIFFSLVFITACSSGLVFVRVTTNPVDLWSAPSSQARLEKEYFDQHFGPFFRTEQLIIRAPLTDKHIYQPYPSGADVPFGPPLDIQILHQVLDLQIAIENITASYDNETVTLQDICLAPLSPYNTNCTILSVLNYFQNSHSVLDHKKGDDFFVYADYHTHFLYCVRFINFVKNYKNPNLTISFTAERSIEDELNRESDSDVFTVVISYAIMFLYISLALGHMKSCRRLLVDSKVSLGIAGILIVLSSVACSLGVFSYIGLPLTLIVIEVIPFLVLAVGVDNIFILVQAYQRDERLQGETLDQQLGRVLGEVAPSMFLSSFSETVAFFLGALSVMPAVHTFSLFAGLAVFIDFLLQITCFVSLLGLDIKRQEKNRLDIFCCVRGAEDGTSVQASESCLFRFFKNSYSPLLLKDWMRPIVIAIFVGVLSFSIAVLNKVDIGLDQSLSMPDDSYMVDYFKSISQYLHAGPPVYFVLEEGHDYTSSKGQNMVCGGMGCNNDSLVQQIFNAAQLDNYTRIGFAPSSWIDDYFDWVKPQSSCCRVDNITDQFCNASVVDPACVRCRPLTPEGKQRPQGGDFMRFLPMFLSDNPNPKCGKGGHAAYSSAVNILLGHGTRVGATYFMTYHTVLQTSADFIDALKKARLIASNVTETMGINGSAYRVFPYSVFYVFYEQYLTIIDDTIFNLGVSLGAIFLVTMVLLGCELWSAVIMCATIAMVLVNMFGVMWLWGISLNAVSLVNLVMSCGISVEFCSHITRAFTVSMKGSRVERAEEALAHMGSSVFSGITLTKFGGIVVLAFAKSQIFQIFYFRMYLAMVLLGATHGLIFLPVLLSYIGPSVNKAKSCATEERYKGTERERLLNF"
with open(DNA_SEQ_FILENAME_TXT , "r") as file: NPC1_GENE_SEQ_Chrom18 = file.read() # Open the file and read its content
mRNA_seq_FILTERED,\
    mRNA_seq_length_Nt,\
        Dictionnary_of_CodonCounts_SORTED,\
            N_Synonymous_Gene_Sequences_NPC1_chromose18,\
                total_combinations_EXPONENTIAL_PART_BASE_10,\
                    AverageDegenerationDegreePerCodon\
                        = CountSynonymousmRNASequences_from_mRNA(NPC1_GENE_SEQ_Chrom18 , FORMAT = gene_sequence_FILE_FORMAT , mRNASeq_StartingWord = DNA_GENE_SEQ_StartingWord , PrintMessage = PrintMessage , ShowCombinationEvolutionPlot = ShowCombinationEvolutionPlot , ORDER_TYPE = ORDER_TYPE)

# Calculating the mRNA molar mass in g/mol
Molar_mass_mRNA = mrna_molar_mass(mRNA_seq_FILTERED)
print(f"\nmRNA molar mass = {Molar_mass_mRNA} g/mol")

# Reading the csv or xlsx file
print("# Reading CSV file")
RAW_DATA = read_xlsx(filename , MUTE_MESSAGES = True)

# Cleaning the empty rows and columns in the csv or xlsx file
print("# Cleaning the empty rows and columns in the CSV file")
NOT_EMPTY_DATA = clean_empty_rows_and_columns(RAW_DATA)

# Extracting the labels, units and numerical values in the CSV file
print("# Extracting the labels, units and numerical values in the CSV file")
List_of_all_datalabels , List_of_all_timelabels , List_of_all_mRNA_dosages , List_of_all_time , List_of_all_protein_concentrations \
    = extracting_time_and_concentration(NOT_EMPTY_DATA,
                                        expected_data_Nb_per_excel_sheet             = Nb_Data_Per_Sheet,
                                        expected_row_for_datalabels                  = Data_Labels_Row,
                                        expected_row_for_timelabels                  = Time_Labels_Row,
                                        expected_row_for_mRNA_dosages                = mRNA_Dosages_Row,
                                        expected_row_for_time_START                  = Time_Row_START,
                                        expected_row_for_protein_concentration_START = Protein_Concentration_Row_START,
                                        expected_column_offset_for_datalabels        = Data_Labels_Column_OFFSET,
                                        expected_column_offset_for_timelabels        = Time_Labels_Column_OFFSET,
                                        expected_column_offset_for_mRNA_dosages      = mRNA_Dosages_Column_OFFSET,
                                        expected_columns_per_data                    = Nb_Columns_Per_Data,
                                        BOOL_ReplaceNonBreakingSpace_by_space        = BOOL_ReplaceNonBreakingSpace_by_space)

# Storing the number of curves to analyse
Nb_curves_to_fit = len(List_of_all_protein_concentrations)

# Normalizing some data
print("# Normalizing some data")
NORM_ProteinConcentrations_ALL_TOGETHER = normalize_array(List_of_all_protein_concentrations , NORM_VAL_MIN , NORM_VAL_MAX)
NORM_ProteinConcentrations_ONE_by_ONE   = np.array([normalize_array(List_of_all_protein_concentrations[array_num] , NORM_VAL_MIN , NORM_VAL_MAX) for array_num in range(Nb_curves_to_fit)])

# Printing the estimated time to compute every set of parameters
print("# Printing the estimated time to compute every set of parameters")
Nb_MSE_to_compute_per_AP                         = Nb_points_per_decade_AP * (1 + (MAX_Power_AP - MIN_Power_AP + 1) // PowerStep_AP)
Nb_MSE_to_compute_per_BP                         = Nb_points_per_decade_BP * (1 + (MAX_Power_BP - MIN_Power_BP + 1) // PowerStep_BP)
Nb_MSE_to_compute_per_Am                         = Nb_points_per_decade_Am * (1 + (MAX_Power_Am - MIN_Power_Am + 1) // PowerStep_Am)
Nb_MSE_to_compute_per_Bm                         = Nb_points_per_decade_Bm * (1 + (MAX_Power_Bm - MIN_Power_Bm + 1) // PowerStep_Bm)
Nb_MSE_to_compute_per_curve                      = Nb_MSE_to_compute_per_AP * Nb_MSE_to_compute_per_BP * Nb_MSE_to_compute_per_Am * Nb_MSE_to_compute_per_Bm
Nb_MSE_to_compute_all_curves                     = Nb_curves_to_fit * Nb_MSE_to_compute_per_curve

# For loops generating the list of all the values that every parameter will take for searching the optimal set of parameters
SearchingRange_alpha_protein    = np.zeros(Nb_MSE_to_compute_per_AP)
SearchingRange_beta_protein     = np.zeros(Nb_MSE_to_compute_per_BP)
SearchingRange_alpha_mRNA       = np.zeros(Nb_MSE_to_compute_per_Am)
SearchingRange_beta_mRNA        = np.zeros(Nb_MSE_to_compute_per_Bm)

count = 0
if METHOD_GENERATE_PARAMETERS_GRID_SEARCH == "Power_min_to_power_max":
    count = 0
    for decade_power_AP in range(MIN_Power_AP , MAX_Power_AP + 1 , PowerStep_AP):
        for value_in_the_decade in np.linspace(1 , 10 , num = Nb_points_per_decade_AP, endpoint = False):
            new_value_for_parameters            = value_in_the_decade * 10.**decade_power_AP
            SearchingRange_alpha_protein[count] = new_value_for_parameters
            count = count + 1
    count = 0
    for decade_power_BP in range(MIN_Power_BP , MAX_Power_BP + 1 , PowerStep_BP):
        for value_in_the_decade in np.linspace(1 , 10 , num = Nb_points_per_decade_BP, endpoint = False):
            new_value_for_parameters            = value_in_the_decade * 10.**decade_power_BP
            SearchingRange_beta_protein[count]  = new_value_for_parameters
            count = count + 1
    count = 0
    for decade_power_Am in range(MIN_Power_Am , MAX_Power_Am + 1 , PowerStep_Am):
        for value_in_the_decade in np.linspace(1 , 10 , num = Nb_points_per_decade_Am, endpoint = False):
            new_value_for_parameters            = value_in_the_decade * 10.**decade_power_Am
            SearchingRange_alpha_mRNA[count]    = new_value_for_parameters
            count = count + 1
    count = 0
    for decade_power_Bm in range(MIN_Power_Bm , MAX_Power_Bm + 1 , PowerStep_Bm):
        for value_in_the_decade in np.linspace(1 , 10 , num = Nb_points_per_decade_Bm, endpoint = False):
            new_value_for_parameters            = value_in_the_decade * 10.**decade_power_Bm
            SearchingRange_beta_mRNA[count]     = new_value_for_parameters
            count = count + 1
elif METHOD_GENERATE_PARAMETERS_GRID_SEARCH == "Powers_in_a_given_list":
    for decade_power in List_of_Power10_values_for_parameters:
        for value_in_the_decade in range(1 , Nb_points_per_decade + 1 , 1):
            new_value_for_parameters            = value_in_the_decade * 10.**decade_power
            SearchingRange_alpha_protein[count] = new_value_for_parameters
            SearchingRange_beta_protein[count]  = new_value_for_parameters
            SearchingRange_alpha_mRNA[count]    = new_value_for_parameters
            SearchingRange_beta_mRNA[count]     = new_value_for_parameters
            count = count + 1

# Ploting the curves
print("# Ploting the curves")
fig_EXPERIMENTAL_DATA = plt.figure(figsize = (DefaultFigureWidth , DefaultFigureHeight) , dpi = DefaultDPIResolution , facecolor = DefaultFaceColor , edgecolor = DefaultEdgeColor , layout = DefaultLayout)
plt.subplot(2 , 2 , (1,2))
plt.title("Protein concentration vs time (varying mRNA dosage)")
for num_curve in range(Nb_curves_to_fit):
    Data_label  = List_of_all_datalabels[num_curve]
    Time_label  = List_of_all_timelabels[num_curve]
    mRNA_dosage = List_of_all_mRNA_dosages[num_curve]
    Ltime       = List_of_all_time[num_curve]
    LProtConc   = List_of_all_protein_concentrations[num_curve]
    color_index_tmp = int((len(Lcolors)-1) * num_curve / (Nb_curves_to_fit - 1)) if Nb_curves_to_fit >= 2 else 0 # this makes the curve color varying between red and black (Red -> Yellow -> Green -> Cyan -> Blue -> Black)
    color_tmp = Lcolors[color_index_tmp]
    label_tmp = f"mRNA dosage = {mRNA_dosage}"
    plt.xlabel("Time [H]")
    plt.ylabel("[Protein](t) [-]")
    plt.plot(Ltime , LProtConc , color = color_tmp , linestyle = DefaultLineStyle , marker = '.' , label = label_tmp)
plt.grid(visible = DefaultGridValue , which = 'both' , alpha = DefaultGridOpacity , color = DefaultGridColor , linestyle = DefaultGridLineStyle , linewidth = DefaultGridLineWidth)
plt.legend()

plt.subplot(2 , 2 , 3)
plt.title("NORMALIZED protein concentration vs time (varying mRNA dosage)\nNormalizing all the curves together")
for num_curve in range(Nb_curves_to_fit):
    Data_label          = List_of_all_datalabels[num_curve]
    Time_label          = List_of_all_timelabels[num_curve]
    mRNA_dosage         = List_of_all_mRNA_dosages[num_curve]
    Ltime               = List_of_all_time[num_curve]
    LProtConc_NORM_ALL  = NORM_ProteinConcentrations_ALL_TOGETHER[num_curve]
    color_index_tmp = int((len(Lcolors)-1) * num_curve / (Nb_curves_to_fit - 1)) if Nb_curves_to_fit >= 2 else 0 # this makes the curve color varying between red and black (Red -> Yellow -> Green -> Cyan -> Blue -> Black)
    color_tmp = Lcolors[color_index_tmp]
    label_tmp = f"mRNA dosage = {mRNA_dosage}"
    plt.xlabel("Time [H]")
    plt.ylabel("[Protein](t) [-]")
    plt.plot(Ltime , LProtConc_NORM_ALL , color = color_tmp , linestyle = DefaultLineStyle , marker = '.' , label = label_tmp)
plt.grid(visible = DefaultGridValue , which = 'both' , alpha = DefaultGridOpacity , color = DefaultGridColor , linestyle = DefaultGridLineStyle , linewidth = DefaultGridLineWidth)
plt.legend()

plt.subplot(2 , 2 , 4)
plt.title("NORMALIZED protein concentration vs time (varying mRNA dosage)\nNormalizing every curve independently")
for num_curve in range(Nb_curves_to_fit):
    Data_label          = List_of_all_datalabels[num_curve]
    Time_label          = List_of_all_timelabels[num_curve]
    mRNA_dosage         = List_of_all_mRNA_dosages[num_curve]
    Ltime               = List_of_all_time[num_curve]
    LProtConc_NORM_1by1 = NORM_ProteinConcentrations_ONE_by_ONE[num_curve]
    color_index_tmp = int((len(Lcolors)-1) * num_curve / (Nb_curves_to_fit - 1)) if Nb_curves_to_fit >= 2 else 0 # this makes the curve color varying between red and black (Red -> Yellow -> Green -> Cyan -> Blue -> Black)
    color_tmp = Lcolors[color_index_tmp]
    label_tmp = f"mRNA dosage = {mRNA_dosage}"
    plt.xlabel("Time [H]")
    plt.ylabel("[Protein](t) [-]")
    plt.plot(Ltime , LProtConc_NORM_1by1 , color = color_tmp , linestyle = DefaultLineStyle , marker = '.' , label = label_tmp)
plt.grid(visible = DefaultGridValue , which = 'both' , alpha = DefaultGridOpacity , color = DefaultGridColor , linestyle = DefaultGridLineStyle , linewidth = DefaultGridLineWidth)
plt.legend()
plt.show()

# Filename prefix
fig_filename_prefix = "EXPERIMENTAL Protein concentration vs time (varying mRNA dosage)"

# Saving the figure containing the training and validation losses curves
save_plot_with_timestamp(directory          = os.path.join(path_to_all_files , 'SAVED_FIGURES_CURVE_FITTING'),
                         figure_name        = fig_EXPERIMENTAL_DATA,
                         filename_prefix    = fig_filename_prefix,
                         timestamp          = DATE_AND_TIME, # Format Year_Month_Day_Hour_Minute_Second
                         Print_message      = True)

# Initializing the lists containing the further optimized parameters
List_of_alpha_protein_optimized = np.zeros(Nb_curves_to_fit)
List_of_beta_protein_optimized  = np.zeros(Nb_curves_to_fit)
List_of_alpha_mRNA_optimized    = np.zeros(Nb_curves_to_fit)
List_of_beta_mRNA_optimized     = np.zeros(Nb_curves_to_fit)
List_of_MSE_all_curves          = []
List_of_Tau_half_life_mRNA      = np.zeros(Nb_curves_to_fit)
List_of_Tau_half_life_protein   = np.zeros(Nb_curves_to_fit)
List_of_mRNA_masses             = np.zeros(Nb_curves_to_fit)
List_of_mRNA_mass_units         = np.zeros(Nb_curves_to_fit)
List_of_mRNA_mass_units_str     = []

######################################################################################################################################################################
## NESTED SEARCHING LOOPS TO DETERMINE THE NUMBER OF MSEs TO BE LATER COMPUTED => THIS STEP IS NECESSARY TO CORRECTLY HANDLE THE MEMORY ALLOCATION FOR THE MSE list ##
print("## NESTED SEARCHING LOOPS TO DETERMINE THE NUMBER OF MSEs TO BE LATER COMPUTED => THIS STEP IS NECESSARY TO CORRECTLY HANDLE THE MEMORY ALLOCATION FOR THE MSE list ##")

# Initializing the list of adjusted number of MSEs to be later computed (every adjusted number of MSE to compute is equal or lower than the theoretical maximum number of MSEs to compute)
List_of_adjusted_Nb_of_MSEs_to_compute = np.zeros(Nb_curves_to_fit)
for num_curve , EXPERIMENTAL_protein_concentration in enumerate(List_of_all_protein_concentrations):
    
    # Measuring the starting time to compute one optimization process
    ComputationalTimeForOneMSE_REAL_t_injection = time.time()
    
    # Initializing the count variable
    count                             = 0
    Nb_MSE_to_compute_IGNORED         = 0
    PreviousReadingPercentage         = 0
    List_of_alpha_protein_potential   = []
    List_of_beta_protein_potential    = []
    List_of_alpha_mRNA_potential      = []
    List_of_beta_mRNA_potential       = []
    LOWEST_MSE                        = -1
    List_of_MSE                       = []
    ADJUSTED_NUMBER_OF_MSE_TO_COMPUTE = 0 # Initializing the adjusted number of MSEs to be later computed to 0
    
    # Message for clear information on the prompt / console
    print('\n' + 0 * '\t' + f"CURVE FITTING {num_curve+1}/{Nb_curves_to_fit} ...\n")
    
    # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
    ## CHECKING WHETHER THE PROTEIN MAXIMUM CONCENTRATION IS PRECEDED AND FOLLOWED BY A SMALLER VALUE => INFLEXION POINT IS NECESSARY FOR THE MODEL
    BOOL_Inflexion_point_in_the_EXPERIMENTAL_data = any(EXPERIMENTAL_protein_concentration[i - 1] < max(EXPERIMENTAL_protein_concentration) and EXPERIMENTAL_protein_concentration[i + 1] < max(EXPERIMENTAL_protein_concentration) for i in range(1, len(EXPERIMENTAL_protein_concentration) - 1) if EXPERIMENTAL_protein_concentration[i] == max(EXPERIMENTAL_protein_concentration))
    if not BOOL_Inflexion_point_in_the_EXPERIMENTAL_data: pass
    else:
        
        # Normalizing the experimental curve between some estimated values in terms of (either) concentration (or) number of protein molecules
        EXPERIMENTAL_protein_concentration_NORMALIZED = normalize_array(EXPERIMENTAL_protein_concentration,
                                                                        DEF_MIN = Number_of_NPC1_proteins_before_injection_at_equilibrium_____ESTIMATED,
                                                                        DEF_MAX = Number_of_NPC1_proteins_before_injection_at_peak_MAX_ESTIMATED)
        
        # Extracting the starting and ending time variables
        List_of_time_in_HOUR   = List_of_all_time[num_curve]
        List_of_time_in_SECOND = List_of_time_in_HOUR * 3600
        t_injection            = List_of_time_in_SECOND[0] # automatically scaling the results on the experimental timeline
        t_final                = List_of_time_in_SECOND[-1] # automatically scaling the results on the experimental timeline
        
        # Storing the index of the value right before and right after the EXPERIMENTAL MAXIMUM VALUE
        index_of_EXPERIMENTAL_MAXIMUM_VALUE = np.argmax(EXPERIMENTAL_protein_concentration)
        index_BEFORE_MAX = index_of_EXPERIMENTAL_MAXIMUM_VALUE - 1 if index_of_EXPERIMENTAL_MAXIMUM_VALUE > 0 else 0
        index_AFTER_MAX  = index_of_EXPERIMENTAL_MAXIMUM_VALUE + 1 if index_of_EXPERIMENTAL_MAXIMUM_VALUE < len(EXPERIMENTAL_protein_concentration) - 1 else index_of_EXPERIMENTAL_MAXIMUM_VALUE
        
        # Storing the times before and after the MAX
        timepoint_BEFORE_MAX_in_SECOND = List_of_time_in_SECOND[index_BEFORE_MAX]
        timepoint_AFTER_MAX_in_SECOND  = List_of_time_in_SECOND[index_AFTER_MAX]
        
        # # # # # # # # # # # # # # # # # # # # # # # # #
        ## EXTRACTING THE mRNA MASS FROM THE DATA LABELS
        mRNA_dosage = List_of_all_mRNA_dosages[num_curve]
        # Regular expression to match a number followed by a space and 'pg'
        sliced_string = re.search(r'(\d+(\.\d+)?)\s*([a-zA-Z]+)', mRNA_dosage)
        if sliced_string : mRNA_mass , mRNA_mass_unit_str = int(sliced_string.group(1)) , sliced_string.group(3) # Convert the matched number to an integer and extract the unit as a string
        else : raise ValueError("No valid mRNA mass found in the mRNA_dosage string.")
        
        # Interprets the unit string to convert it into a numerical value
        if   'kg' in mRNA_mass_unit_str : mRNA_mass_unit = 1e3
        elif 'dg' in mRNA_mass_unit_str : mRNA_mass_unit = 1e-1
        elif 'cg' in mRNA_mass_unit_str : mRNA_mass_unit = 1e-2
        elif 'mg' in mRNA_mass_unit_str : mRNA_mass_unit = 1e-3
        elif 'µg' in mRNA_mass_unit_str : mRNA_mass_unit = 1e-6
        elif 'ng' in mRNA_mass_unit_str : mRNA_mass_unit = 1e-9
        elif 'pg' in mRNA_mass_unit_str : mRNA_mass_unit = 1e-12
        elif 'g'  in mRNA_mass_unit_str : mRNA_mass_unit = 1 # the gram is the standard unit because the molar masses are in g/mol
        else : raise ValueError("No interpretable unit found in the mRNA dosage => mRNA mass unit is unknown ...")
        
        # Calculating the current mRNA mass
        mRNA_mass_in_kg                    = mRNA_mass * mRNA_mass_unit
        mRNA_mass_in_g                     = mRNA_mass_in_kg * 1000
        List_of_mRNA_masses[num_curve]     = mRNA_mass
        List_of_mRNA_mass_units[num_curve] = mRNA_mass_unit
        List_of_mRNA_mass_units_str.append(mRNA_mass_unit_str)
        
        # Calculating the current number of mRNA molecules
        mRNA_t_injection = mRNA_mass_in_g * Avogadro / Molar_mass_mRNA # The number of mRNA at t = t_injection (no unit as it is a number of molecules)
        
        # # # # # # # # # # # # # # # # # #
        ## BEGINING OF THE NESTED FOR LOOPS
        # Engulfing the next messages under this 'text banner'
        print("-------------------------------------------------------------------------------------------------------------------------------------------------")
        print("| alpha_P | beta_P | alpha_mRNA | beta_mRNA | alpha P/mRNA | beta P/mRNA | alpha_P/beta_P | alpha_mRNA/beta_mRNA | Maximum Theor. protein conc. |")
        print("-------------------------------------------------------------------------------------------------------------------------------------------------")
        
        # For loop on the beta_protein parameter
        for i_B_prot , beta_protein in enumerate(SearchingRange_beta_protein):
            
            if not(beta_protein < 1): Nb_MSE_to_compute_IGNORED += len(SearchingRange_beta_mRNA) * len(SearchingRange_alpha_protein) * len(SearchingRange_alpha_mRNA)
            else:
                
                # For loop on the beta_mRNA parameter
                for i_B_mRNA , beta_mRNA in enumerate(SearchingRange_beta_mRNA):
                    
                    # Calculate the inflexion timepoint in seconds
                    t_inflexion = calculate_protein_theoretical_inflexion_timepoint(t_injection, beta_protein, beta_mRNA)
                    BOOL_inflexion_is_between_previous_and_next_timepoint_around_EXPERIMENTAL_MAXIMUM_VALUE = (timepoint_BEFORE_MAX_in_SECOND <= t_inflexion) and (t_inflexion <= timepoint_AFTER_MAX_in_SECOND)
                    
                    if not(beta_protein < beta_mRNA and beta_mRNA * beta_protein < 1) or not(BOOL_inflexion_is_between_previous_and_next_timepoint_around_EXPERIMENTAL_MAXIMUM_VALUE): Nb_MSE_to_compute_IGNORED += len(SearchingRange_alpha_protein) * len(SearchingRange_alpha_mRNA)
                    else:
                        
                        # For loop on the alpha_protein parameter
                        for i_A_prot , alpha_protein in enumerate(SearchingRange_alpha_protein):
                            
                            # For loop on the alpha_mRNA parameter
                            for i_A_mRNA , alpha_mRNA in enumerate(SearchingRange_alpha_mRNA):
                                
                                # Calculate the protein concentration at equilibrium
                                protein_eq = alpha_protein * alpha_mRNA / (beta_protein * beta_mRNA)
                                
                                # Calculating the boolean telling whether the predicted concentration of proteins at equilibrium is in the expected range or not ...
                                BOOL_protein_eq_predicted_is_in_the_expected_range = (Number_of_NPC1_proteins_before_injection_at_equilibrium_MIN_ESTIMATED <= protein_eq) and (protein_eq <= Number_of_NPC1_proteins_before_injection_at_equilibrium_MAX_ESTIMATED)
                                
                                # Calculate the maximum protein concentration predicted by the model
                                protein_concentration_THEOR_MAX = calculate_theoretical_maximum_concentration(alpha_protein, beta_protein, alpha_mRNA, beta_mRNA, mRNA_t_injection)
                                BOOL_maximum_protein_concentration_is_in_expected_range = (protein_eq < protein_concentration_THEOR_MAX) and (protein_concentration_THEOR_MAX <= coef_multiplicator_between_1_and_2_for_estimated_MAX * Number_of_NPC1_proteins_before_injection_at_peak_MAX_ESTIMATED)
                                
                                if not(protein_eq >= 1 and BOOL_protein_eq_predicted_is_in_the_expected_range): Nb_MSE_to_compute_IGNORED = Nb_MSE_to_compute_IGNORED + 1
                                else:
                                    
                                    # Showing the progression of the optimization process
                                    PreviousReadingPercentage = Show_Progression(TIME_START_LOOP = ComputationalTimeForOneMSE_REAL_t_injection, # time start in seconds
                                                     TIME_NOW                              = time.time(), # current time in seconds
                                                     Number_of_tabulations_before_sentence = 1,
                                                     DescriptiveWord                       = f'\tCurve fitting {num_curve+1}/{Nb_curves_to_fit}',
                                                     current_index                         = count + Nb_MSE_to_compute_IGNORED,
                                                     index_start                           = 0,
                                                     index_stop                            = (Nb_MSE_to_compute_all_curves - 1)//Nb_curves_to_fit,
                                                     PreviousReadingPercentage             = PreviousReadingPercentage,
                                                     PercentageStep                        = ShowProgression_PercentageStep,
                                                     round_digit                           = round_digit_ShowProgression,
                                                     BOOL_do_not_show_time_left            = False)
                                    count = count + 1
                                    
                                    # Calculating the current mean square error
                                    THEOR_protein_concentration_EXP_scale = Generate_List_of_protein_t_AFTER_INJECTION(alpha_protein, beta_protein, List_of_time_in_SECOND, t_injection, alpha_mRNA, beta_mRNA, mRNA_t_injection)
                                    MSE = calculate_MSE(THEOR_protein_concentration_EXP_scale, EXPERIMENTAL_protein_concentration_NORMALIZED)
                                    
                                    # Compare to the lowest MSE value to directly save the optimized parameters
                                    if MSE < LOWEST_MSE or LOWEST_MSE == -1:
                                        
                                        # Refreshing and adding the lowest MSE value to the list
                                        LOWEST_MSE = MSE                                        
                                        List_of_MSE.append(LOWEST_MSE)
                                        
                                        # Adding the parameters found to their respective lists
                                        List_of_alpha_protein_potential.append(alpha_protein)
                                        List_of_beta_protein_potential.append(beta_protein)
                                        List_of_alpha_mRNA_potential.append(alpha_mRNA)
                                        List_of_beta_mRNA_potential.append(beta_mRNA)
                                        
                                        ratio_AP_AM = round(math.log10(alpha_protein / alpha_mRNA),2)
                                        ratio_Ap_Bp = round(math.log10(alpha_protein / beta_protein),2)
                                        ratio_BP_BM = round(math.log10(beta_protein / beta_mRNA),2)
                                        ratio_Am_Bm = round(math.log10(alpha_mRNA / beta_mRNA),2)
                                        
                                        # Printing the main message
                                        message_tmp = f"| e{round(math.log10(alpha_protein),2)} \t| e{round(math.log10(beta_protein),2)} \t| e{round(math.log10(alpha_mRNA),2)}   \t| e{round(math.log10(beta_mRNA),2)} \t| e{ratio_AP_AM}  \t| e{ratio_BP_BM} \t| e{ratio_Ap_Bp}  \t| e{ratio_Am_Bm}  \t| e{round(math.log10(protein_concentration_THEOR_MAX), 2)}  \t|"
                                        print(f"{message_tmp}")
                                        
                                        # Plotting the curves only on the condition that this boolean is True
                                        if BOOL_plot_intermediate_curves_with_EXP_DATA_and_FIT:
                                            
                                            # Calculating the current mean square error
                                            List_of_time_expanded_tmp = np.linspace(t_injection, t_final, num = Nb_points_per_hour_THEORETICAL_CURVE, endpoint = True)
                                            FIT_tmp = Generate_List_of_protein_t_AFTER_INJECTION(alpha_protein, beta_protein, List_of_time_expanded_tmp, t_injection, alpha_mRNA, beta_mRNA, mRNA_t_injection)
                                            
                                            fig_tmp = plt.figure(figsize = (DefaultFigureWidth , DefaultFigureHeight) , dpi = DefaultDPIResolution , facecolor = DefaultFaceColor , edgecolor = DefaultEdgeColor , layout = DefaultLayout)
                                            plt.title(message_tmp, fontsize = DefaultFontSize)
                                            plt.plot(List_of_time_in_HOUR, EXPERIMENTAL_protein_concentration_NORMALIZED, color = ReturnRGBColorCode('red'), label = 'protein_exp(t)')
                                            plt.plot(List_of_time_expanded_tmp/3600, FIT_tmp, color = ReturnRGBColorCode('magenta'), label = 'protein_FIT(t)')
                                            plt.xlim([min(List_of_time_in_HOUR), max(List_of_time_in_HOUR)])
                                            # # Set log scale for Y-axis
                                            # plt.yscale('log')
                                            # Adds the rest of the figure settings
                                            plt.xlabel('time (h)')
                                            plt.ylabel('[protein](t)')
                                            plt.legend(fontsize = DefaultFontSize , loc = 'upper right')
                                            plt.grid(visible = DefaultGridValue , which = 'both' , alpha = DefaultGridOpacity , color = DefaultGridColor , linestyle = DefaultGridLineStyle , linewidth = DefaultGridLineWidth)
                                            plt.show()
    
    # Storing the current list of MSE in the 'all MSE' list
    Nb_of_MSEs_to_plot = len(List_of_MSE)
    List_of_MSE_all_curves.append(List_of_MSE)
    X_axis = np.cumsum(np.ones(Nb_of_MSEs_to_plot))
    
    if Nb_of_MSEs_to_plot == 0:
        print('WARNING : 0 MSE were calculated due to 0 set of parameters satisfying the constraints / conditions ...')
        List_of_Tau_half_life_mRNA[num_curve]    = -1
        List_of_Tau_half_life_protein[num_curve] = -1
    else:
        
        ###################################
        ## STORING THE RESULTS IN THE LISTS
        List_of_alpha_protein_optimized[num_curve]  = List_of_alpha_protein_potential[-1]
        List_of_beta_protein_optimized[num_curve]   = List_of_beta_protein_potential[-1]
        List_of_alpha_mRNA_optimized[num_curve]     = List_of_alpha_mRNA_potential[-1]
        List_of_beta_mRNA_optimized[num_curve]      = List_of_beta_mRNA_potential[-1]
        
        # Storing the optimal parameters found
        alpha_protein_OPTI = List_of_alpha_protein_optimized[num_curve]
        beta_protein_OPTI  = List_of_beta_protein_optimized[num_curve]
        alpha_mRNA_OPTI    = List_of_alpha_mRNA_optimized[num_curve]
        beta_mRNA_OPTI     = List_of_beta_mRNA_optimized[num_curve]
        
        # Decomposing the optimal parameters in integer and exponential part
        alpha_protein_OPTI_int , alpha_protein_OPTI_exp = extract_value_and_exponent(alpha_protein_OPTI)
        beta_protein_OPTI_int  , beta_protein_OPTI_exp  = extract_value_and_exponent(beta_protein_OPTI)
        alpha_mRNA_OPTI_int    , alpha_mRNA_OPTI_exp    = extract_value_and_exponent(alpha_mRNA_OPTI)
        beta_mRNA_OPTI_int     , beta_mRNA_OPTI_exp     = extract_value_and_exponent(beta_mRNA_OPTI)
        
        ##########################################
        ## CALCULATE THE TAU HALF-LIFE OF THE mRNA
        [t_injection_plus_Tau_half_life_mRNA, Tau_half_life_mRNA, mRNA_at_Tau_half_life_mRNA] = find_Tau_half_life_mRNA_from_optimal_parameters(t_injection, alpha_mRNA_OPTI, beta_mRNA_OPTI, mRNA_t_injection)
        t_injection_plus_Tau_half_life_mRNA_in_HOUR = t_injection_plus_Tau_half_life_mRNA / 3600
        Tau_half_life_mRNA_in_HOUR                  = round(Tau_half_life_mRNA / 3600 , 2)
        List_of_Tau_half_life_mRNA[num_curve]       = Tau_half_life_mRNA_in_HOUR
        
        #############################################
        ## CALCULATE THE TAU HALF-LIFE OF THE PROTEIN
        MAXIMUM_TIMEPOINT = coef_t_limit_Tau_half_life_protein_finding * t_final
        [t_inflexion, P_at_t_inflexion, t_injection_plus_Tau_half_life_protein_in_SEC, P_at_Tau_half_life_protein] = find_Tau_half_life_protein_from_optimal_parameters(alpha_protein_OPTI, beta_protein_OPTI, t_injection, alpha_mRNA_OPTI, beta_mRNA_OPTI, mRNA_t_injection, Number_of_iterations_MAX, MAXIMUM_TIMEPOINT, precision_required_as_decimal_value)
        t_injection_plus_Tau_half_life_protein_in_HOUR = t_injection_plus_Tau_half_life_protein_in_SEC / 3600
        Tau_half_life_protein_in_SEC                   = t_injection_plus_Tau_half_life_protein_in_SEC - t_injection
        Tau_half_life_protein_in_HOUR                  = round(Tau_half_life_protein_in_SEC / 3600 , 2)
        List_of_Tau_half_life_protein[num_curve]       = Tau_half_life_protein_in_HOUR
        
        #####################
        ## PRINT SOME RESULTS
        print(f"\n\t alpha_protein_OPTI      = {alpha_protein_OPTI_int} * 10^{alpha_protein_OPTI_exp}")
        print(f"\t beta_protein_OPTI       = {beta_protein_OPTI_int} * 10^{beta_protein_OPTI_exp}")
        print(f"\t alpha_mRNA_OPTI         = {alpha_mRNA_OPTI_int} * 10^{alpha_mRNA_OPTI_exp}")
        print(f"\t beta_mRNA_OPTI          = {beta_mRNA_OPTI_int} * 10^{beta_mRNA_OPTI_exp}")
        print(f"\t LOWEST_MSE              = {LOWEST_MSE}")
        print(f"\t Tau HALF-LIFE (mRNA)    = {Tau_half_life_mRNA_in_HOUR} hours")
        print(f"\t Tau HALF-LIFE (protein) = {Tau_half_life_protein_in_HOUR} hours")
        
        ############################################
        ## GENERATING THE FIT CURVE WITH MORE POINTS
        List_of_time_expanded_SEC  = np.linspace(t_injection, MAXIMUM_TIMEPOINT, num = Nb_points_per_hour_THEORETICAL_CURVE, endpoint = True)
        List_of_time_expanded_HOUR = List_of_time_expanded_SEC /3600
        OPTI_mRNA_concentration    = Generate_List_of_mRNA_t_AFTER_INJECTION(alpha_mRNA_OPTI, beta_mRNA_OPTI, mRNA_t_injection, List_of_time_expanded_SEC, t_injection)
        OPTI_protein_concentration = Generate_List_of_protein_t_AFTER_INJECTION(alpha_protein_OPTI, beta_protein_OPTI, List_of_time_expanded_SEC, t_injection, alpha_mRNA_OPTI, beta_mRNA_OPTI, mRNA_t_injection)
        
        ######################
        ## Plotting the curves
        fig_mRNA_and_protein_over_time = plt.figure(figsize = (DefaultFigureWidth , DefaultFigureHeight) , dpi = DefaultDPIResolution , facecolor = DefaultFaceColor , edgecolor = DefaultEdgeColor , layout = DefaultLayout)
        
        #################
        # # SUBPLOT 1 # #
        plt.subplot(2 , 2 , 1)
        plt.title(f"mRNA concentration over time - FIT\nmRNA dosage = {mRNA_dosage}")
        plt.plot(List_of_time_expanded_HOUR, OPTI_mRNA_concentration, color = ReturnRGBColorCode('gray'), label = '[mRNA](t) (FIT)')
        # Plot a horizontal line at mRNA concentration = 'mRNA_at_Tau_half_life'
        plt.axhline(y = mRNA_at_Tau_half_life_mRNA, color = ReturnRGBColorCode('gray'), linestyle = '--', label = '[mRNA](t=t_injection + Tau mRNA half-life)')
        # Plot a vertical line at t = t_injection_plus_Tau_half_life_mRNA_in_HOUR
        label_tmp = f"t_injection + Tau mRNA half-life = {round(t_injection_plus_Tau_half_life_mRNA_in_HOUR,2)}h\n(Tau half-life = {Tau_half_life_mRNA_in_HOUR}h)"
        plt.axvline(x = t_injection_plus_Tau_half_life_mRNA_in_HOUR, color = ReturnRGBColorCode('gray'), linestyle = '--', label = label_tmp)
        plt.xlim([min(List_of_time_expanded_HOUR), max(max(List_of_time_expanded_HOUR), t_injection_plus_Tau_half_life_mRNA_in_HOUR + 10)])
        # # Set log scale for Y-axis
        # plt.yscale('log')
        # Adds the rest of the figure settings
        plt.xlabel('time (h)')
        plt.ylabel('[mRNA](t)')
        plt.legend(fontsize = DefaultFontSize , loc = 'upper right')
        plt.grid(visible = DefaultGridValue , which = 'both' , alpha = DefaultGridOpacity , color = DefaultGridColor , linestyle = DefaultGridLineStyle , linewidth = DefaultGridLineWidth)
        
        #################
        # # SUBPLOT 2 # #
        plt.subplot(2 , 2 , 2)
        plt.title(f"Protein concentration over time - FIT\nmRNA dosage = {mRNA_dosage}")
        plt.plot(List_of_time_in_HOUR, EXPERIMENTAL_protein_concentration_NORMALIZED, color = ReturnRGBColorCode('red'), label = '[protein](t) (EXP)')
        plt.plot(List_of_time_expanded_HOUR, OPTI_protein_concentration, color = ReturnRGBColorCode('gray'), label = '[protein](t) (FIT)')
        # Plot a horizontal line at protein concentration = 'P_at_Tau_half_life_protein'
        plt.axhline(y = P_at_Tau_half_life_protein, color = ReturnRGBColorCode('gray'), linestyle = '--', label = '[protein](t=t_injection + Tau protein half-life)')
        # Plot a vertical line at t = t_injection_plus_Tau_half_life_protein_in_SEC
        label_tmp = f"t_injection + Tau protein half-life = {round(t_injection_plus_Tau_half_life_protein_in_HOUR,2)}h\n(Tau half-life = {Tau_half_life_protein_in_HOUR}h)"
        plt.axvline(x = t_injection_plus_Tau_half_life_protein_in_HOUR, color = ReturnRGBColorCode('gray'), linestyle = '--', label = label_tmp)
        plt.xlim([min(List_of_time_in_HOUR), max(max(List_of_time_in_HOUR), t_injection_plus_Tau_half_life_protein_in_HOUR + 10)])
        # # Set log scale for Y-axis
        # plt.yscale('log')
        # Adds the rest of the figure settings
        plt.xlabel('time (h)')
        plt.ylabel('[protein](t)')
        plt.legend(fontsize = DefaultFontSize , loc = 'upper right')
        plt.grid(visible = DefaultGridValue , which = 'both' , alpha = DefaultGridOpacity , color = DefaultGridColor , linestyle = DefaultGridLineStyle , linewidth = DefaultGridLineWidth)
        
        #################
        # # SUBPLOT 3 # #
        plt.subplot(2 , 2 , 3)
        plt.title("MSE over iterations - log scale")
        plt.plot(X_axis, List_of_MSE, color = ReturnRGBColorCode('magenta')  , label = 'MSE')
        plt.xlim([0, max(X_axis)])
        # Set log scale for Y-axis
        plt.yscale('log')
        # Adds the rest of the figure settings
        plt.xlabel('iterations')
        plt.ylabel('MSE (LOG SCALE)')
        plt.legend(fontsize = DefaultFontSize , loc = 'upper right')
        plt.grid(visible = DefaultGridValue , which = 'both' , alpha = DefaultGridOpacity , color = DefaultGridColor , linestyle = DefaultGridLineStyle , linewidth = DefaultGridLineWidth)
        
        #################
        # # SUBPLOT 4 # #
        plt.subplot(2 , 2 , 4)
        plt.title("Parameters evolution over iterations - log scale")
        plt.plot(X_axis, List_of_alpha_protein_potential, color = ReturnRGBColorCode('indigo'), label = 'alpha_P [1/s]')
        plt.plot(X_axis, List_of_beta_protein_potential, color = ReturnRGBColorCode('orange'), label = 'beta_P [1/s]')
        plt.plot(X_axis, List_of_alpha_mRNA_potential, color = ReturnRGBColorCode('blue'), label = 'alpha_M [1/s]')
        plt.plot(X_axis, List_of_beta_mRNA_potential, color = ReturnRGBColorCode('red'), label = 'beta_M [1/s]')
        plt.xlim([0, max(X_axis)])
        # Set log scale for Y-axis
        plt.yscale('log')
        # Adds the rest of the figure settings
        plt.xlabel('iterations')
        plt.ylabel('Parameters')
        plt.legend(fontsize = DefaultFontSize , loc = 'upper right')
        plt.grid(visible = DefaultGridValue , which = 'both' , alpha = DefaultGridOpacity , color = DefaultGridColor , linestyle = DefaultGridLineStyle , linewidth = DefaultGridLineWidth)
        
        plt.show()
        
        # Automatically including the hyperparameters in the filename
        fig_filename_prefix = f"OPTI_params_curve_fitting_{num_curve}"
        
        # Saving the figure containing the training and validation losses curves
        save_plot_with_timestamp(directory          = os.path.join(path_to_all_files , 'SAVED_FIGURES_CURVE_FITTING'),
                                 figure_name        = fig_mRNA_and_protein_over_time,
                                 filename_prefix    = fig_filename_prefix,
                                 timestamp          = DATE_AND_TIME, # Format Year_Month_Day_Hour_Minute_Second
                                 Print_message      = True)
    
    List_of_adjusted_Nb_of_MSEs_to_compute[num_curve] = count
    print(f"Curve fitting {num_curve+1}/{Nb_curves_to_fit}: Number of MSE to compute = {count}")

########################################################################
## Plotting the evolution of Tau half-life versus the mRNA injected mass
min_mRNA_mass_unit_in_g = min(List_of_mRNA_mass_units)
List_of_mRNA_masses_scaled_on_lowest_unit = List_of_mRNA_masses * List_of_mRNA_mass_units / min_mRNA_mass_unit_in_g

customized_label_mRNA = 'Tau_mRNA(mRNA injected)'
customized_label_protein = 'Tau_Protein(mRNA injected)'
for i , Tau_protein_tmp in enumerate(List_of_Tau_half_life_protein):
    Tau_mRNA_tmp = List_of_Tau_half_life_mRNA[i]
    
    mRNA_mass_units_str_tmp  = List_of_mRNA_mass_units_str[i]
    customized_label_mRNA    = customized_label_mRNA    + f'\n - Tau({i}) = {round(Tau_mRNA_tmp,2)}h ({List_of_mRNA_masses[i]} {mRNA_mass_units_str_tmp})'
    customized_label_protein = customized_label_protein + f'\n - Tau({i}) = {round(Tau_protein_tmp,2)}h ({List_of_mRNA_masses[i]} {mRNA_mass_units_str_tmp})'

fig_Tau_half_life_mRNA_and_protein_over_mRNA_mass_injected = plt.figure(figsize = (DefaultFigureWidth , DefaultFigureHeight) , dpi = DefaultDPIResolution , facecolor = DefaultFaceColor , edgecolor = DefaultEdgeColor , layout = DefaultLayout)
plt.title("Tau Half-Life over mRNA mass injected (per cell)")
plt.plot(List_of_mRNA_masses_scaled_on_lowest_unit, List_of_Tau_half_life_mRNA, color = ReturnRGBColorCode('red'), label = customized_label_mRNA)
plt.plot(List_of_mRNA_masses_scaled_on_lowest_unit, List_of_Tau_half_life_protein, color = ReturnRGBColorCode('blue'), label = customized_label_protein)
# plt.xlim([min(List_of_mRNA_masses_scaled_on_lowest_unit), max(List_of_mRNA_masses_scaled_on_lowest_unit)])
plt.xlabel(f'mRNA mass injected [{min_mRNA_mass_unit_in_g}g]')
plt.ylabel('Tau half-life [h]')
plt.legend(fontsize = DefaultFontSize , loc = 'upper right')
plt.grid(visible = DefaultGridValue , which = 'both' , alpha = DefaultGridOpacity , color = DefaultGridColor , linestyle = DefaultGridLineStyle , linewidth = DefaultGridLineWidth)
plt.show()

# Filename prefix
fig_filename_prefix = "Tau_half_life_protein_over_mRNA_mass_injected"

# Saving the figure containing the training and validation losses curves
save_plot_with_timestamp(directory          = os.path.join(path_to_all_files , 'SAVED_FIGURES_CURVE_FITTING'),
                         figure_name        = fig_Tau_half_life_mRNA_and_protein_over_mRNA_mass_injected,
                         filename_prefix    = fig_filename_prefix,
                         timestamp          = DATE_AND_TIME, # Format Year_Month_Day_Hour_Minute_Second
                         Print_message      = True)

#################
## END OF PROGRAM
TIME_END_PROGRAM = time.time()
PrintElapsedTime(TIME_START_PROGRAM , TIME_END_PROGRAM)